#LIL MILAN CLASSES:

import pandas as pd
import numpy as np
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# filepath = "C:\\Users\\isabe\\Downloads\\"

def run(uploaded_ordini_files, uploaded_other_files, mese, anno):
    print("Starting run function")
    
    #ordini
    ordini_processor = Ordini(uploaded_ordini_files, mese = mese)
    ordini = ordini_processor.preprocess()

    try:
        #run matchers
        shopify_matcher = ShopifyMatcher(uploaded_other_files, df_ordini=ordini)
        scalapay_matcher = ScalapayMatcher(uploaded_other_files, df_ordini=ordini)
        satispay_matcher = SatispayMatcher(uploaded_other_files, df_ordini=ordini)
        paypal_matcher = PaypalMatcher(uploaded_other_files, df_ordini=ordini)
        qromo_matcher = QromoMatcher(uploaded_other_files, df_ordini=ordini)
        bonifico_matcher = BonificoMatcher(uploaded_other_files, df_ordini=ordini)

        # Create the matchers list
        matchers = [
            shopify_matcher,
            scalapay_matcher,
            satispay_matcher,
            paypal_matcher,
            qromo_matcher,
            bonifico_matcher
        ]

       #excel
        runner = MatcherRunner(matchers, ordini)
        print("Runner created")
        
        result = runner.run_all_matchers(mese, anno)
        print(f"Result type: {type(result)}")
        print(f"Result info:", result.info() if hasattr(result, 'info') else "No info")
        
        return result
    except Exception as e:
        print(f"Error in run: {str(e)}")
        raise e


    # #excel
    # runner = MatcherRunner(matchers, ordini)
    # final_df = runner.run_all_matchers(mese, anno)
    # # df_check_all, df_ordini_all, df = runner.run_all_matchers(mese=mese, anno= anno)


#CLASSE CHE PRENDE IN ENTRATA FILE DI ORDINI, LO PULISCE E LO RESTITUISCE PULITO
class Ordini:
    def __init__(self, uploaded_files, mese):
        """
        uploaded_files: A dictionary where keys are file names and values are UploadedFile objects
        mese: The month of interest
        """
        self.uploaded_files = uploaded_files  # Dictionary of UploadedFile objects
        self.df = None  # This will hold the DataFrame once loaded
        self.mese = mese

    def load_data(self):
    
        lil_file = self.uploaded_files.get("Ordini LIL")
        agee_file = self.uploaded_files.get("Ordini AGEE")
        
        # assicurarsi che esista
        if lil_file:
            lil = pd.read_csv(lil_file, dtype={'Lineitem sku': 'string', 'Device ID': 'string', 'Id': 'string', "Tags": "string", "Next Payment Due At": "string"})
            lil["Brand"] = "LIL Milan"
        else:
            lil = pd.DataFrame()  # or handle the missing file as needed

        # assicurarsi che esista
        if agee_file:
            agee = pd.read_csv(agee_file, dtype={'Lineitem sku': 'string', 'Device ID': 'string', 'Id': 'string', "Tags": "string", "Next Payment Due At": "string"})
            agee["Brand"] = "AGEE"
        else:
            agee = pd.DataFrame()  # or handle the missing file as needed

        # Concatenate dataframes if both are available
        self.df = pd.concat([lil, agee], ignore_index=True) if len(lil) > 0 or len(agee) > 0 else pd.DataFrame()

# class Ordini:
#     def __init__(self, filepath, mese):
#         self.filepath = filepath
#         self.df = None  # This will hold the DataFrame once loaded
#         self.mese = mese

#     def load_data(self):
#         if self.mese == 8:
#         # Load the CSV data
#             self.df = pd.read_excel(self.filepath + "Vendite Agosto v3.xlsx", sheet_name="Generale LIL")
#             self.df = self.df[self.df["Name"].str.startswith("#")]
#             self.df["Brand"] = "LIL Milan"
#         elif self.mese == 9:
#             lil = pd.read_csv(self.filepath + "Ordini LIL.csv", dtype={'Lineitem sku': 'string', 'Device ID': 'string', 'Id': 'string', "Tags": "string", "Next Payment Due At": "string"})
#             agee = pd.read_csv(self.filepath + "Ordini AGEE 9.csv", dtype={'Lineitem sku': 'string', 'Device ID': 'string', 'Id': 'string', "Tags": "string", "Next Payment Due At": "string"})
#             lil["Brand"] = "LIL Milan"
#             agee["Brand"] = "AGEE"
#             self.df = pd.concat([lil, agee], ignore_index=True)
#         elif self.mese == 10:
#             self.df = pd.read_csv(self.filepath + "ordini2.csv", dtype={'Lineitem sku': 'string', 'Device ID': 'string', 'Id': 'string', "Tags": "string", "Next Payment Due At": "string"})
#             self.df["Brand"] = "LIL Milan"
        
        # Forward-fill NaN values for the same Name
        self.df[self.df.columns] = self.df.groupby('Name')[self.df.columns].ffill()

    def adjust_financial_status(self):
        # PARTIALLY_PAID handling
        partially_paid_names = self.df[self.df["Financial Status"] == "partially_paid"]["Name"].unique()
        nomi_cambi = self.df.loc[(self.df['Lineitem compare at price'] == 0) & 
                                 (self.df['Lineitem price'] != 0) & 
                                 (self.df['Total'] != 0), 'Name'].unique()
        mask = self.df["Name"].isin(partially_paid_names) & ~self.df["Name"].isin(nomi_cambi)

        for name in self.df.loc[mask, "Name"].unique():
            name_mask = self.df["Name"] == name
            new_total = self.df.loc[name_mask, "Total"].values[0] - self.df.loc[name_mask, "Outstanding Balance"].values[0]
    
            if new_total >= 0:
                self.df.loc[name_mask, "Total"] = new_total

        
        # REFUNDED handling
        refunded_names = self.df[self.df["Financial Status"] == "refunded"]["Name"].unique()
        for name in refunded_names:
            mask = self.df["Name"] == name
            total_value = self.df.loc[mask, "Total"].iloc[0]
            if total_value != 0:
                diff = self.df.loc[mask, "Total"] - self.df.loc[mask, "Refunded Amount"]
                self.df.loc[mask & (diff == 0), "Total"] = 0
                self.df.loc[mask & (diff == 0), "Lineitem quantity"] = 0

    def handle_discounts(self):
        # Handle discount codes (e.g., Roma100, MILANO100%, etc.)
        # sconto100_names = self.df[(self.df["Discount Code"].str.contains('100', case=False, na=False)) & 
        #                           (self.df["Total"] == 0)]["Name"].unique()
        self.df["Payment Method"] = self.df["Payment Method"].str.strip()
        sconto100_names = self.df[(self.df["Payment Method"].isna()) & 
                                  (self.df["Total"] == 0)]["Name"].unique()
        for name in sconto100_names:
            mask = self.df["Name"] == name
            self.df.loc[mask, "CHECK"] = "SCONTO100"

    def adjust_country(self):        
        # Handle missing Shipping Country
        self.df.loc[((self.df["Shipping Method"] == "Firgun House") | 
                     (self.df["Shipping Method"].isna())) & 
                    (self.df["Shipping Country"].isna()), "Shipping Country"] = "IT"

    def handle_gift_card(self):
        # Handle gift card payments
        gift_card_names = self.df[self.df["Payment Method"] == "Gift Card"]["Name"].unique()
        mask = self.df["Name"].isin(gift_card_names)
        self.df.loc[mask, "CHECK"] = "FALSO"
        self.df.loc[mask, "CHECK_VALORE_GIFT_CARD"] = self.df.loc[mask, "Total"]

    def normalize_lengths_payments(self, row):
        
        # Ensure methods and references are lists
        methods = row['Payment Method List'] if isinstance(row['Payment Method List'], list) else []
        references = row['Payment Reference List'] if isinstance(row['Payment Reference List'], list) else []
        
        # Initialize the paired list and placeholder
        paired = []
        first_method = methods[0] if methods else np.nan
        
        # Pair methods and references, filling missing values as needed
        for i in range(max(len(methods), len(references))):
            method = methods[i] if i < len(methods) else first_method
            reference = references[i] if i < len(references) else np.nan
            paired.append((method, reference))
        
        return paired

    def handle_payments(self):
        # #Handle payment types
        self.df["Payment Method"] = self.df["Payment Method"].str.replace("Custom (POS)", "Qromo")
        self.df["Payment Method"] = self.df["Payment Method"].str.replace("custom|Wire Transfer", "Bonifico", regex=True)

    def handle_location(self):
        # Add LIL House London Locations
        london_locations = self.df.loc[self.df['Location'].isna() & (self.df['Shipping Method'] == "Standard")]
        self.df.loc[london_locations.index, 'Location'] = "LIL House London"
        
        # Fill missing Location values with default value
        location_nan = self.df.groupby('Name')['Location'].transform(lambda x: x.isna().all())
        self.df.loc[location_nan, 'Location'] = "Firgun House"

        self.df['Location'] = self.df['Location'].str.replace(r'(?i)\blil house\b', 'LIL House', regex=True)

    def apply_cambi(self):
        # Apply the changes defined in the 'ordini_con_cambi' logic
        self.df = self.ordini_con_cambi()

    def ordini_con_cambi(self):
        nomi_cambi = self.df.loc[(self.df['Lineitem compare at price'] == 0) & 
                                 (self.df['Lineitem price'] != 0) & 
                                 (self.df['Total'] != 0), 'Name'].unique()

        # Process each group by 'Name' for the return handling logic
        for name, group in self.df.groupby('Name'):
            if name in nomi_cambi:
                items_comprati_dopo = group[group['Lineitem compare at price'] == 0]
                primi_items = group[group['Lineitem compare at price'].isna()]
                primi_items_gioielli = primi_items[primi_items['Lineitem price'] > 10]

                total = group['Total'].values[0]
                shipping = group['Shipping'].values[0]
                discount = group['Discount Amount'].values[0]
            
                #Solo un oggetto è stato scambiato
                if len(items_comprati_dopo) == 1: 

                    item_comprato_dopo_price = items_comprati_dopo['Lineitem price'].values[0]
                    item_comprato_dopo_quantity = items_comprati_dopo['Lineitem quantity'].values[0]
                    
                    #solo un oggetto tra quelli comprati inizialmente è candidato per essere stato scambiato (supponiamo che il cliente non scambia borse, scatole ed engraving)
                    if len(primi_items_gioielli) == 1 and item_comprato_dopo_price > 10: #se è minore o uguale a 10, l'ha probabilmente aggiunto

                        item_restituito_price = primi_items_gioielli['Lineitem price'].values[0]
                        item_restituito_quantity = primi_items_gioielli['Lineitem quantity'].values[0]
                    
                        items_tenuti_prices = (primi_items.loc[~primi_items.index.isin(primi_items_gioielli.index)].apply(lambda row: row['Lineitem price'] * row['Lineitem quantity'], axis=1)
                                            .sum()) + (item_comprato_dopo_price * item_comprato_dopo_quantity) #primi items acquistati che non sono gioielli + item comprato dopo
                        
                        new_total = total - (item_restituito_price * item_comprato_dopo_quantity)  #tolgo tante unità dell'oggetto comprato prima quanti oggeti sono stati comprati dopo
                        check_total = items_tenuti_prices + shipping - discount + item_restituito_price * (item_restituito_quantity - item_comprato_dopo_quantity)
                        
                        if new_total == check_total and new_total > 0:
                            self.df.loc[group.index, 'Total'] = new_total
                            self.df.loc[primi_items_gioielli.index, 'Lineitem quantity'] = item_restituito_quantity - item_comprato_dopo_quantity

                    #più di un oggetto tra quelli comprati inizialmente potrebbe essere stato scambiato
                    elif len(primi_items_gioielli) > 1 and item_comprato_dopo_price > 10:

                        matched_row = primi_items_gioielli[primi_items_gioielli['Lineitem price'] == item_comprato_dopo_price]

                        # Check if there's a match and retrieve its price and quantity
                        if len(matched_row) == 1:
                            item_restituito_price = matched_row['Lineitem price'].values[0]
                            item_restituito_quantity = matched_row['Lineitem quantity'].values[0]

                            items_tenuti_prices = (primi_items.loc[~primi_items.index.isin(matched_row.index)].apply(lambda row: row['Lineitem price'] * row['Lineitem quantity'], axis=1)
                                            .sum()) + (item_comprato_dopo_price * item_comprato_dopo_quantity) #primi items acquistati che non sono gioielli + item comprato dopo
                        
                            new_total = total - (item_restituito_price * item_comprato_dopo_quantity)  #tolgo tante unità dell'oggetto comprato prima quanti oggeti sono stati comprati dopo
                            check_total = items_tenuti_prices + shipping - discount + item_restituito_price * (item_restituito_quantity - item_comprato_dopo_quantity)

                            if new_total == check_total and new_total > 0:
                                self.df.loc[group.index, 'Total'] = new_total
                                self.df.loc[matched_row.index, 'Lineitem quantity'] = item_restituito_quantity - item_comprato_dopo_quantity

                
                #più oggeti sono stati scambiati
                elif len(items_comprati_dopo) == 2:
                    
                    #più oggetti consecutivamente
                    if len(primi_items_gioielli) == 1: #si assume che avvengano due resi consecutivi

                        ultimo_item_tenuto = items_comprati_dopo.iloc[-1] #si assume siano in ordine
                        ultimo_item_tenuto_price = ultimo_item_tenuto['Lineitem price']
                        ultimo_item_tenuto_quantity = ultimo_item_tenuto['Lineitem quantity']

                        item_aggiunto_e_reso = items_comprati_dopo.loc[items_comprati_dopo.index != ultimo_item_tenuto.name]
                        item_aggiunto_e_reso_quantity = item_aggiunto_e_reso['Lineitem quantity'].values[0]

                        item_restituito_price = primi_items_gioielli['Lineitem price'].values[0]
                        item_restituito_quantity = primi_items_gioielli['Lineitem quantity'].values[0]

                        if ultimo_item_tenuto_quantity == item_aggiunto_e_reso_quantity == item_restituito_quantity:
                            primi_items = pd.concat([primi_items, item_aggiunto_e_reso])
                            tutti_gioielli_tranne_ultimo = primi_items[primi_items['Lineitem price'] > 10]

                            items_tenuti_prices = (primi_items.loc[~primi_items.index.isin(tutti_gioielli_tranne_ultimo.index)].apply(lambda row: row['Lineitem price'] * row['Lineitem quantity'], axis=1)
                                                .sum()) + (ultimo_item_tenuto_price * ultimo_item_tenuto_quantity) #primi items acquistati che non sono gioielli + item comprato dopo
                            
                            items_restituiti_prices = (tutti_gioielli_tranne_ultimo.apply(lambda row: row['Lineitem price'] * row['Lineitem quantity'], axis=1).sum())
                        
                            new_total = total - items_restituiti_prices
                            check_total = items_tenuti_prices + shipping - discount 

                            if new_total == check_total and new_total > 0:
                                self.df.loc[group.index, 'Total'] = new_total
                                self.df.loc[item_aggiunto_e_reso.index, 'Lineitem quantity'] = 0
                                self.df.loc[primi_items_gioielli.index, 'Lineitem quantity'] = 0

        return self.df

    def preprocess(self):
        # Call all preprocessing steps in sequence
        self.load_data()
        self.adjust_financial_status()
        self.handle_discounts()
        self.adjust_country()
        self.handle_gift_card()
        self.handle_payments()
        self.handle_location()
        self.apply_cambi()

        return self.df
    

#PAGAMENTI: PRENDE IN ENTRATA I VARI METODI DI PAGAMENTO, GLI ORDINI PULITI E TROVA I MATCH. 
#RESTITUISCE IL DF DEI MATCH CHE SERVE PER MOSTRARE I CASI DA CONTROLLARE, IL DF DEGLI ORDINI ULTERIORMENTE CONTROLLATI E IL DF DEI PAGAMENTI CON AGGIUNTA COLONNA CHECK
class PaymentMatcher:
    payment_info_list = [] 

    def __init__(self, uploaded_files, df_ordini):
        self.uploaded_files = uploaded_files
        self.df_ordini = df_ordini

    #fare check se va controllato
    def check_values(self, row, tipo):
        total = row["Total"]
        amount = self.get_amount(row, tipo)

        if pd.isna(total) or pd.isna(amount):
            return "NON TROVATO"
        elif abs(total - amount) <= 0.05:
            return "VERO"
        else:
            return "FALSO"

    #Calcolare valore teorico della gift card
    def calculate_gift_card(self, row, tipo):
        amount = self.get_amount(row, tipo)
        if row["CHECK"] == "FALSO" and "Gift Card" in str(row["Payment Method"]):
            return row["Total"] - amount
        elif row["CHECK"] == "NON TROVATO" and "Gift Card" in str(row["Payment Method"]):
            return row["Total"] - amount
        else:
            return np.nan

    #controlla se il discount è giusto
    def check_discounts(self, row, df_check, tipo):
        
        nomi = df_check.loc[(df_check["CHECK"] == "FALSO") & 
                        (df_check["CHECK_VALORE_GIFT_CARD"].isna()) & 
                        (df_check["Lineitem compare at price"] == 0)]["Name"]
    
        # If nomi is empty, return df_check unchanged
        if len(nomi) == 0:
            return df_check
        
        for name, group in df_check.groupby('Name'):
            if name in nomi.values:
                items_comprati_dopo = group[group['Lineitem compare at price'] == 0]
                primi_items = group[group['Lineitem compare at price'].isna()]
                primi_items_gioielli = primi_items[primi_items['Lineitem price'] > 10]

                total = group['Total'].values[0]
                shipping = group['Shipping'].values[0]
                discount = group['Discount Amount'].values[0]

                amount = self.get_amount(row, tipo)

                # Solo un oggetto è stato scambiato
                if len(items_comprati_dopo) == 1: 

                    item_comprato_dopo_price = items_comprati_dopo['Lineitem price'].values[0]
                    item_comprato_dopo_quantity = items_comprati_dopo['Lineitem quantity'].values[0]

                    if len(primi_items_gioielli) == 1 and item_comprato_dopo_price > 10:

                        item_restituito_price = primi_items_gioielli['Lineitem price'].values[0]
                        item_restituito_quantity = primi_items_gioielli['Lineitem quantity'].values[0]

                        check_discount = discount / (item_restituito_price * (item_restituito_quantity + item_comprato_dopo_quantity) + item_comprato_dopo_price * item_comprato_dopo_quantity) 
                        calculated_discount = check_discount * item_comprato_dopo_price * item_comprato_dopo_quantity

                        items_tenuti_prices = (
                            primi_items.loc[~primi_items.index.isin(primi_items_gioielli.index)]
                            .apply(lambda row: row['Lineitem price'] * row['Lineitem quantity'], axis=1)
                            .sum()) + (item_comprato_dopo_price * item_comprato_dopo_quantity)
                        
                        new_total = total + discount - calculated_discount
                        check_total = items_tenuti_prices + shipping - calculated_discount

                        if new_total == check_total and new_total == amount:
                            # Modify df_check instead of self.df_ordini
                            df_check.loc[group.index, 'Total'] = new_total
                            df_check.loc[group.index, 'Discount Amount'] = calculated_discount
                            df_check.loc[group.index, 'CHECK'] = "VERO"

                    elif len(primi_items_gioielli) > 1 and item_comprato_dopo_price > 10:

                        matched_row = primi_items_gioielli[primi_items_gioielli['Lineitem price'] == item_comprato_dopo_price]

                        if len(matched_row) == 1:
                            item_restituito_price = matched_row['Lineitem price'].values[0]
                            item_restituito_quantity = matched_row['Lineitem quantity'].values[0]

                            check_discount = discount / (item_restituito_price * (item_restituito_quantity + item_comprato_dopo_quantity) + item_comprato_dopo_price * item_comprato_dopo_quantity) 
                            calculated_discount = check_discount * item_comprato_dopo_price * item_comprato_dopo_quantity

                            items_tenuti_prices = (
                                primi_items.loc[~primi_items.index.isin(matched_row.index)]
                                .apply(lambda row: row['Lineitem price'] * row['Lineitem quantity'], axis=1)
                                .sum()) + (item_comprato_dopo_price * item_comprato_dopo_quantity)
                            
                            new_total = total + discount - calculated_discount
                            check_total = items_tenuti_prices + shipping - calculated_discount

                            if new_total == check_total and new_total == amount:
                                df_check.loc[group.index, 'Total'] = new_total
                                df_check.loc[group.index, 'Discount Amount'] = calculated_discount
                                df_check.loc[group.index, 'CHECK'] = "VERO"
                    
                #più oggeti sono stati scambiati
                elif len(items_comprati_dopo) == 2:
                    
                    #più oggetti consecutivamente
                    if len(primi_items_gioielli) == 1: #si assume che avvengano due resi consecutivi

                        ultimo_item_tenuto = items_comprati_dopo.iloc[-1] #si assume siano in ordine
                        ultimo_item_tenuto_price = ultimo_item_tenuto['Lineitem price']
                        ultimo_item_tenuto_quantity = ultimo_item_tenuto['Lineitem quantity']

                        item_aggiunto_e_reso = items_comprati_dopo.loc[items_comprati_dopo.index != ultimo_item_tenuto.name]
                
                        item_restituito_price = primi_items_gioielli['Lineitem price'].values[0]
                        item_restituito_quantity = primi_items_gioielli['Lineitem quantity'].values[0]
                        
                        primi_items = pd.concat([primi_items, item_aggiunto_e_reso])
                        tutti_gioielli_tranne_ultimo = primi_items[primi_items['Lineitem price'] > 10]

                        check_discount = discount / (tutti_gioielli_tranne_ultimo.apply(lambda row: row['Lineitem price'] * (row['Lineitem quantity'] + ultimo_item_tenuto_quantity), axis=1).sum()
                                                    + (ultimo_item_tenuto_price * ultimo_item_tenuto_quantity))
                        calculated_discount = check_discount * ultimo_item_tenuto_price * ultimo_item_tenuto_quantity

                        items_tenuti_prices = (primi_items.loc[~primi_items.index.isin(tutti_gioielli_tranne_ultimo.index)].apply(lambda row: row['Lineitem price'] * row['Lineitem quantity'], axis=1)
                                            .sum()) + (ultimo_item_tenuto_price * ultimo_item_tenuto_quantity) #primi items acquistati che non sono gioielli + item comprato dopo
                        
                        # items_restituiti_prices = (tutti_gioielli_tranne_ultimo.apply(lambda row: row['Lineitem price'] * row['Lineitem quantity'], axis=1).sum())
                    
                        new_total = total + discount - calculated_discount
                        check_total = items_tenuti_prices + shipping - calculated_discount

                        if new_total == check_total and new_total == amount:
                            df_check.loc[group.index, 'Total'] = new_total
                            df_check.loc[group.index, 'Discount Amount'] = calculated_discount
                            df_check.loc[group.index, 'CHECK'] = "VERO"

        return df_check

    # PARTIALLY_REFUNDED handling
    def check_partially_refunded(self, df, amount_column):
        partially_refunded_names = df[(df["Financial Status"] == "partially_refunded") & (df["CHECK"] == "FALSO")]["Name"].unique()
        partially_paid_names = df[(df["Financial Status"] == "partially_paid")  & (df["CHECK"] == "FALSO")]["Name"].unique()
        
        cambi_names = df.loc[(df['Lineitem compare at price'] == 0) & 
                                 (df['Lineitem price'] != 0) & 
                                 (df['Total'] != 0), 'Name'].unique()
        
        mask = df["Name"].isin(partially_refunded_names) & ~df["Name"].isin(cambi_names)
        for name in df.loc[mask, "Name"].unique():
            name_mask = df["Name"] == name
            amount = df.loc[name_mask, amount_column].iloc[0]
            new_total = df.loc[name_mask, "Total"].iloc[0] - df.loc[name_mask, "Refunded Amount"].iloc[0]
            if new_total == amount:
                df.loc[name_mask, "Total"] = new_total
                df.loc[name_mask, "CHECK"] = "VERO"

        mask = df["Name"].isin(partially_paid_names) & df["Name"].isin(cambi_names)
        for name in df.loc[mask, "Name"].unique():
            name_mask = df["Name"] == name
            amount = df.loc[mask, amount_column].iloc[0]
            new_total = df.loc[name_mask, "Total"].iloc[0] - df.loc[name_mask, "Outstanding Balance"].iloc[0]
            if new_total == amount:
                df.loc[name_mask, "Total"] = new_total
                df.loc[mask, "CHECK"] = "VERO"
                             
        return df
    
    #in base al metodo di pagamento, la colonna dell'amount differisce                 
    def get_amount(self, row, tipo):
        """To be implemented in subclass, depending on the 'tipo'"""
        raise NotImplementedError

    #applica tutti i check
    def apply_checks(self, df_check, tipo):
        df_check["CHECK"] = df_check.apply(lambda row: self.check_values(row, tipo), axis=1)
        df_check["CHECK_VALORE_GIFT_CARD"] = df_check.apply(lambda row: self.calculate_gift_card(row, tipo), axis=1)

        cols = df_check.columns.tolist()
        cols = [col for col in cols if col not in ["CHECK", "CHECK_VALORE_GIFT_CARD"]] + ["CHECK", "CHECK_VALORE_GIFT_CARD"]
        df_check = df_check[cols]
    
        return df_check

    #gestisce pagamenti con più di un payment reference
    def match_references(self, payment_reference, merchant_ids):
    # Split the Payment Reference if there is a "+" sign
        references = payment_reference.split('+') if '+' in payment_reference else [payment_reference]
        
        # Initialize a list to hold all matching references
        matching_references = []
        
        # Check if any of the references matches a Merchant ID
        for ref in references:
            stripped_ref = ref.strip()  # Strip whitespace
            if stripped_ref in merchant_ids:
                matching_references.append(stripped_ref)  # Add matching reference to the list
        
        return matching_references  # Return the list of matching references
    

    def merge_dfs(self, ordini, pagamenti, id_column):
        merchant_ids = pagamenti[id_column].astype(str).str.strip().tolist()
        ordini['Matched Reference'] = ordini['Payment References'].apply(lambda ref: self.match_references(ref, merchant_ids))
        ordini = ordini.explode('Matched Reference')
        ordini.reset_index(drop=True, inplace=True)

        check = pd.merge(ordini, pagamenti, left_on="Matched Reference", right_on=id_column, how='outer')

        return ordini, check

    #check if some orders have been paid with 2 different payments
    def check_double_payments(self, df, amount_column):
        # Step 1: Filter the DataFrame for 'FALSO' in 'CHECK'
        filtered_check = df[df['CHECK'] == 'FALSO']
        grouped = filtered_check.groupby('Name').filter(lambda x: x['Matched Reference'].nunique() > 1)

        aggregated = grouped.groupby(['Name', 'Matched Reference'], as_index=False).agg(
            Amount =(amount_column, 'first'), 
            Total = ("Total", "first")  # Get the first 'Lordo' value for each unique reference
        )

        # Now, sum the unique Lordo values by Name
        final_aggregation = aggregated.groupby('Name').agg(
            Matched_References=('Matched Reference', 'unique'),
            Sum=("Amount", 'sum'),
            Total=('Total', 'first')  # Assuming 'Total' is a column in df
        ).reset_index()

        # Step 4: Check if the sum of Lordo equals Total and update CHECK in df
        for _, row in final_aggregation.iterrows():
            if row['Sum'] == row['Total']:
                # Update CHECK for all rows with this Name in df
                df.loc[df['Name'] == row['Name'], 'CHECK'] = 'VERO'

        return df

#skippare classe se non ci sono pagamenti
class SkipMatcherException(Exception):
    pass

#matcher di paypal
class PaypalMatcher(PaymentMatcher):
    def get_amount(self, row, tipo):
        return row["Lordo"]


    def match(self):

        get_valute = {
            "USD": 0.919548,
            "GBP": 1.172707,
            "CHF": 1.046209,
            "SEK": 0.087655,
            "DKK": 0.134034,
            "HUF": 0.002563,
            "CZK": 0.040004,
            "JPY": 0.006126,
            "NOK": 0.086106, 
            "PLN": 0.231678, 
            "TRY": 0.028939,
            "SGD": 0.68849, 
            "NZD": 0.560897, 
            "HKD": 0.117784,
            "CAD": 0.675812, 
            "AUD": 0.609449,
            "ILS": 0.247803
        }

        paypal_file = self.uploaded_files.get("Paypal")
        
        if not paypal_file:
            raise SkipMatcherException("Non ci sono pagamenti con Paypal")
        
        # Process the file and proceed with matching
        df_full = pd.read_csv(paypal_file)

        df_full['Lordo'] = df_full['Lordo'].str.replace('.', '', regex=False)  # Remove periods (thousands separator)
        df_full['Lordo'] = df_full['Lordo'].str.replace(',', '.', regex=False)  # Replace commas with periods (decimal separator)
        df_full['Lordo'] = pd.to_numeric(df_full['Lordo'], errors='coerce')  # Convert to numeric, coercing errors to NaN        
        df_full = df_full[df_full["Tipo"].isin(["Pagamento Express Checkout", "Rimborso di pagamento"])]
        df_full = df_full[~df_full["Nome"].str.contains("propac", case=False, na=False)] #ha detto di toglierlo

        df = df_full[['Data',"Nome", "Tipo", 'Valuta', 'Lordo', 'N° ordine commerciante', "Titolo oggetto"]]
        df = df.groupby('N° ordine commerciante', as_index=False).agg({'Lordo': 'sum',        # Sum the 'Lordo' values
                                                                               'Valuta': 'first'      # Take the first 'Valuta' value for each group
                                                                               })

        df_ordini = self.df_ordini[self.df_ordini['Payment Method'].str.contains('Paypal', case=False, na=False)]

        df_ordini, df_check = self.merge_dfs(df_ordini, df, 'N° ordine commerciante')
       
        df_check = self.apply_checks(df_check, "paypal")
        df_check = self.check_partially_refunded(df_check, "Lordo")

        df_check["Euro"] = 0.0 
        for index, row in df_check.iterrows():
            if row["CHECK"] == "FALSO" and row["Valuta"] != "EUR":
                row["Euro"] = row["Lordo"] * get_valute[row["Valuta"]]
                if (row["Total"] - 10) < row["Euro"] < (row["Total"] + 10):
                    df_check.at[index, "Lordo"] = row["Euro"] 
                    df_check.at[index, "CHECK"] = "VERO" 
        df_check = df_check.drop("Euro", axis = 1)
    
        # Apply check_discounts function directly for each row and modify df_check
        for _, row in df_check.iterrows():
            df_check = self.check_discounts(row, df_check, "paypal")

        df_check = self.check_double_payments(df_check, "Lordo")

        df_check["Payment Method"] = df_check["Payment Method"].astype(str)
        mask = (df_check["Payment Method"].str.contains(r'\+') 
                & (df_check["CHECK"] == "VERO"))
        df_check.loc[mask & df_check["Payment Method"].str.contains("PayPal Express Checkout"), "Payment Method"] = "PayPal Express Checkout" 

        df_ordini = df_check[self.df_ordini.columns]

        df_full = pd.merge(df_full, df_check[["N° ordine commerciante", "CHECK", "Brand"]], on = ["N° ordine commerciante"], how = "left")
        df_full = df_full.drop_duplicates()

        return df_check, df_ordini, df_full
    
#matcher del bonifico
class BonificoMatcher(PaymentMatcher):
    def get_amount(self, row, tipo):
        return row["Importo"]
    
    def find_header_row(self, excel_file, column_name):

        df_sample = pd.read_excel(excel_file)  # Adjust number as needed
        header_row = 0

        # Find the row containing the known column name
        for idx, row in df_sample.iterrows():
            if column_name in row.values:
                header_row = idx + 1
                break

        df = pd.read_excel(excel_file, header=header_row)
                
        return df
    
    def match(self):
        
        operations_patterns = [
            r'stripe',
            r'paypal',
            r'satispay',
            r'scalapay',
            r'rinascente',
            r'paesi ue',
            r'retail group',
            r'sportello automatico',
        ]
        
        bonifici_file = self.uploaded_files.get("Bonifici")
        
        if not bonifici_file:
            raise SkipMatcherException("Non ci sono bonifici")
        
        # Process the file and proceed with matching
        df_full = self.find_header_row(bonifici_file, "Importo")
        
        mask = ~df_full['Operazione'].str.contains('|'.join(operations_patterns), case=False, regex=True, na=False)
        df_full = df_full[mask]        
        df_full["Data_datetime"] = pd.to_datetime(df_full['Data']).dt.tz_localize(None)

        df_ordini = self.df_ordini[self.df_ordini['Payment Method'].str.contains('Bonifico', case=True, na=False)]
        df_ordini['Paid_datetime'] = pd.to_datetime(df_ordini['Paid at']).dt.tz_localize(None)

        df_check = pd.merge(df_ordini, df_full, left_on='Total', right_on='Importo', how='outer')

        df_check = self.apply_checks(df_check, "bonifico")

        df_check['Days_difference'] = ((df_check['Data_datetime'] - df_check['Paid_datetime']).dt.days).abs() 
    
        # Using .values for direct comparison without index dependency
        mask = df_check['Days_difference'].values == df_check.groupby('Name')['Days_difference'].transform('min').values
        df_check = df_check[mask | df_check['Days_difference'].isna()]

        df_check = self.check_partially_refunded(df_check, "Importo")
        # Apply check_discounts function directly for each row and modify df_check
        for _, row in df_check.iterrows():
            df_check = self.check_discounts(row, df_check, "bonifico")

        df_check["Payment Method"] = df_check["Payment Method"].astype(str)
        mask = (df_check["Payment Method"].str.contains(r'\+') 
                & (df_check["CHECK"] == "VERO"))
        df_check.loc[mask & df_check["Payment Method"].str.contains("Bonifico"), "Payment Method"] = "Bonifico"

        df_ordini = df_check[self.df_ordini.columns]

        df_full = pd.merge(df_full, df_check[["Data_datetime", "CHECK", "Brand"]], on = ["Data_datetime"], how = "left")
        df_full = df_full.drop_duplicates()

        return df_check, df_ordini, df_full

#matcher di qromo
class QromoMatcher(PaymentMatcher):
    def get_amount(self, row, tipo):
        return row["Importo Effettivo"]

    def match(self, mese, anno):

        qromo_file = self.uploaded_files.get("Qromo")
        
        if not qromo_file:
            raise SkipMatcherException("Non ci sono pagamenti col POS")
        
        # Process the file and proceed with matching
        df_full = pd.read_csv(qromo_file, thousands='.')
        data_interesse = f"{anno}-{mese:02}"

        df_full = df_full[df_full["Data"].str.startswith(data_interesse)]
        df_full = df_full[df_full["Stato"] != "Annullato"]
        df_full["Data_datetime"] = pd.to_datetime(df_full["Data"]).dt.tz_localize(None)

        df_full["Importo €"] = df_full["Importo €"].astype(float)
        df_full["Importo rimborsato €"] = df_full["Importo rimborsato €"].astype(float)
        df_full["Importo Effettivo"] = df_full["Importo €"] - df_full["Importo rimborsato €"]
        df_full = df_full[df_full["Importo Effettivo"] != 0]

        df = df_full.copy()
        df['partial_date'] = df['Data_datetime'].dt.date
        df = df[["Stato", "Importo €", "Importo rimborsato €", "Importo Effettivo", "Data_datetime", "partial_date"]]
        df = df.sort_values('Data_datetime')

        df_ordini = self.df_ordini[self.df_ordini['Payment Method'].str.contains('Qromo', case=False, na=False)]
        df_ordini['Paid_datetime'] = pd.to_datetime(df_ordini['Paid at']).dt.tz_localize(None)
        df_ordini['partial_date'] = df_ordini['Paid_datetime'].dt.date
        df_ordini = df_ordini[~df_ordini["Name"].isin(PaymentMatcher.payment_info_list)]
        df_ordini = df_ordini.sort_values('Paid_datetime')

        df_check = pd.merge(df_ordini, df, on='partial_date', how='outer', suffixes=('_qromo', '_ordini'))
        df_check = df_check.drop(columns=['partial_date'])

        df_check = self.apply_checks(df_check, "qromo")

        filtered_df = df_check[df_check['CHECK'] != "NON TROVATO"]
        filtered_df['Time_difference'] = (filtered_df['Data_datetime'] - filtered_df['Paid_datetime']).abs() 
        min_indices = filtered_df.groupby(['Name', "Lineitem name", 'CHECK'])['Time_difference'].idxmin()
        df_min_time_diff = df_check.loc[min_indices]

        df_non_trovato = df_check[df_check['CHECK'] == "NON TROVATO"]
        df_check = pd.concat([df_min_time_diff, df_non_trovato], ignore_index=True)

        names_with_vero = df_check[df_check['CHECK'] == 'VERO']['Name'].unique()
        df_check = df_check[~((df_check['CHECK'] == 'FALSO') & (df_check['Name'].isin(names_with_vero)))]

        df_check = self.check_partially_refunded(df_check, "Importo Effettivo")
        for _, row in df_check.iterrows():
            df_check = self.check_discounts(row, df_check, "qromo")

        # Create a mask for rows that contain '+' in the 'Payment Method', exclude 'Gift Card', and have 'CHECK' set to 'VERO'
        mask = (df_check["Payment Method"].str.contains(r'\+') &
                (df_check["CHECK"] == "VERO"))
    
        df_check.loc[mask & df_check["Payment Method"].str.contains('Qromo'), "Payment Method"] = "Qromo"

        df_ordini = df_check[self.df_ordini.columns]

        df_full = pd.merge(df_full, df_check[["Data_datetime", "CHECK", "Brand"]], on = ["Data_datetime"], how = "left")
        df_full = df_full.drop("Data_datetime", axis = 1)
        df_full["CHECK"] = df_full["CHECK"].fillna("NON TROVATO")
        df_full = df_full.drop_duplicates()

        return df_check, df_ordini, df_full


#matcher di shopify   
class ShopifyMatcher(PaymentMatcher):
    def get_amount(self, row, tipo):
        return row["Amount"]

    def match(self):  

        shopify_lil_file = self.uploaded_files.get("Shopify LIL")
        shopify_agee_file = self.uploaded_files.get("Shopify AGEE")

        # assicurarsi che esista
        if shopify_lil_file:
            lil = pd.read_csv(shopify_lil_file)
        else:
            lil = pd.DataFrame()  # or handle the missing file as needed

        # assicurarsi che esista
        if shopify_agee_file:
            agee = pd.read_csv(shopify_agee_file)
        else:
            agee = pd.DataFrame()  # or handle the missing file as needed

        # Concatenate dataframes if both are available
        df_full = pd.concat([lil, agee], ignore_index=True) if len(lil) > 0 or len(agee) > 0 else pd.DataFrame()
        
        if len(df_full) == 0:
            raise SkipMatcherException("Non ci sono pagamenti con Shopify")
        
        df = df_full.groupby('Order', as_index=False)['Amount'].sum()

        df_ordini = self.df_ordini[self.df_ordini['Payment Method'].str.contains('Shopify', case=False, na=False)]

        df_check = pd.merge(df_ordini, df, left_on="Name", right_on="Order", how='outer')
        df_check = self.apply_checks(df_check, "shopify")
        df_check = self.check_partially_refunded(df_check, "Amount")
    
        # Apply check_discounts function directly for each row and modify df_check
        for _, row in df_check.iterrows():
            df_check = self.check_discounts(row, df_check, "shopify")
        
        df_check["Payment Method"] = df_check["Payment Method"].astype(str)
        mask = (df_check["Payment Method"].str.contains(r'\+') &
                (df_check["CHECK"] == "VERO"))

        df_check.loc[mask & df_check["Payment Method"].str.contains("Shopify Payments"), "Payment Method"] = "Shopify Payments"
        
        df_ordini = df_check[self.df_ordini.columns]

        df_full = pd.merge(df_full, df_check[["Name", "CHECK", "Brand"]], left_on = "Order", right_on = "Name", how = "left")
        df_full = df_full.drop("Name", axis = 1)
        df_full = df_full.drop_duplicates()

        return df_check, df_ordini, df_full
    
#matcher di scalapay
class ScalapayMatcher(PaymentMatcher):
    def get_amount(self, row, tipo):
        return row["Import lordo"]

    def match(self):
        scalapay_file = self.uploaded_files.get("Scalapay")
        
        if not scalapay_file:
            raise SkipMatcherException("Non ci sono pagamenti con Scalapay")
        
        # Process the file and proceed with matching
        df_full = pd.read_csv(scalapay_file)
        
        # df = df_full[["Merchant ID", "Tipo", "Data acquisto/rimborso", "Import lordo"]]
        df = df_full.groupby('Merchant ID', as_index=False, dropna=False)['Import lordo'].sum()

        df_ordini = self.df_ordini[self.df_ordini['Payment Method'].str.contains('Scalapay', case=False, na=False)]

        # Check orders with double reference
        df_ordini, df_check = self.merge_dfs(df_ordini, df, "Merchant ID")

        df_check = self.apply_checks(df_check, "scalapay")
        df_check = self.check_partially_refunded(df_check, "Import lordo")
    
        # Apply check_discounts function directly for each row and modify df_check
        for _, row in df_check.iterrows():
            df_check = self.check_discounts(row, df_check, "scalapay")

        df_check = self.check_double_payments(df_check, "Import lordo")

        df_check["Payment Method"] = df_check["Payment Method"].astype(str)
        mask = (df_check["Payment Method"].str.contains(r'\+') &
                (df_check["CHECK"] == "VERO"))

        df_check.loc[mask & df_check["Payment Method"].str.contains("Scalapay"), "Payment Method"] = "Scalapay"

        df_ordini = df_check[self.df_ordini.columns]

        df_full = pd.merge(df_full, df_check[["Merchant ID", "CHECK", "Brand"]], on = ["Merchant ID"], how = "left")
        df_full = df_full.drop_duplicates()

        return df_check, df_ordini, df_full
    
#matcher di satispay
class SatispayMatcher(PaymentMatcher):
    def get_amount(self, row, tipo):
        return row["total_amount"]
    
    def match(self):

        satispay_file = self.uploaded_files.get("Satispay")
        
        if not satispay_file:
            raise SkipMatcherException("Non ci sono pagamenti con Satispay")
        
        # Process the file and proceed with matching
        df_full = pd.read_csv(satispay_file)
        
        df = df_full[['payment_date', 'total_amount', 'description']]
        df["Data_datetime"] = pd.to_datetime(df["payment_date"] ).dt.tz_localize(None)
        df['partial_date'] = df['Data_datetime'].dt.date

        #Oridni online con description != 0
        df_ordini_online = self.df_ordini[self.df_ordini['Payment Method'].str.contains('Satispay', case=False, na=False)]
        df_ordini_online, df_check_online = self.merge_dfs(df_ordini_online, df[df["description"] != "0"], 'description')

        df_check_online = self.apply_checks(df_check_online, "satispay")
        df_check_online = self.check_partially_refunded(df_check_online, "total_amount")

        for _, row in df_check_online.iterrows():
            df_check_online = self.check_discounts(row, df_check_online, "satispay")

        df_check_online = self.check_double_payments(df_check_online, "total_amount")

        #Ordini negozio con description = 0
        df_ordini_negozio = self.df_ordini[self.df_ordini['Payment Method'].str.contains('Qromo', case=False, na=False)]
        df_ordini_negozio['Paid_datetime'] = pd.to_datetime(df_ordini_negozio['Paid at']).dt.tz_localize(None)
        df_ordini_negozio['partial_date'] = df_ordini_negozio['Paid_datetime'].dt.date

        df_check_negozio = pd.merge(df_ordini_negozio, df[df["description"] == "0"], on="partial_date", how='right')
        df_check_negozio['Time_difference'] = (df_check_negozio['Paid_datetime'] - df_check_negozio['Data_datetime']).abs()    

        df_check_negozio = self.apply_checks(df_check_negozio, "satispay")
        df_check_negozio = self.check_partially_refunded(df_check_negozio, "total_amount")

        for _, row in df_check_negozio.iterrows():
            df_check_negozio = self.check_discounts(row, df_check_negozio, "satispay")

        df_check_negozio = df_check_negozio.loc[df_check_negozio.groupby(['CHECK', "Data_datetime"])['Time_difference'].idxmin()]
        
        names_with_vero = df_check_negozio[(df_check_negozio['CHECK'] == 'VERO')]['Data_datetime'].unique()
        df_check_negozio = df_check_negozio[~((df_check_negozio['CHECK'] == 'FALSO') & (df_check_negozio['Data_datetime'].isin(names_with_vero)))]

        # Store payment_date and total_amount of rows with "VERO"
        for n in names_with_vero:
            PaymentMatcher.payment_info_list.append(n)

        df_check = pd.concat([df_check_online, df_check_negozio])
    
        df_check["Payment Method"] = df_check["Payment Method"].astype(str)
        mask = (df_check["Payment Method"].str.contains(r'\+') &
                (df_check["CHECK"] == "VERO"))

        df_check.loc[mask & df_check["Payment Method"].str.contains("Satispay"), "Payment Method"] = "Satispay"

        df_ordini = df_check[self.df_ordini.columns]

        df_full = pd.merge(df_full, df_check[["payment_date", "description", "CHECK", "Brand"]], on = ["payment_date", "description"], how = "left")
        df_full = df_full.drop_duplicates()

        return df_check, df_ordini, df_full
    


#CLASSE PER RUNNARE I MATCHERS E CREARE EXCEL
class MatcherRunner:
    def __init__(self, matchers, df_ordini_iniziale):
        self.matchers = matchers
        self.df_ordini_iniziale = df_ordini_iniziale
        self.df_ordini_all = None

    #runna matchers e crea excel
    # def run_all_matchers(self, mese, anno=2024):
    #     qromo_matcher = next((matcher for matcher in self.matchers if isinstance(matcher, QromoMatcher)), None)
    #     all_dfs = []
    #     check_dfs = []
    #     self.filename = f'Check_{mese:02}_lilmilan.xlsx'
        
    #     # Create dictionaries to store DataFrames for each sheet
    #     lil_sheets = {}
    #     agee_sheets = {}

    #     for matcher in self.matchers:
    #         try:
    #             if matcher != qromo_matcher:
    #                 df_check, df_ordini, df = matcher.match()
    #             else:
    #                 df_check, df_ordini, df = matcher.match(mese, anno)
    #         except SkipMatcherException as e:
    #             print(f"{e}")
    #             continue

    #         # Append to all_dfs and check_dfs
    #         all_dfs.append(df_ordini)
    #         check_dfs.append(df_check)

    #         # Create masks
    #         mask_lil = df["Brand"] == "LIL Milan"
    #         mask_agee = df["Brand"] == "AGEE"

    #         # Store filtered DataFrames in dictionaries
    #         if mask_lil.any():
    #             payment_name_lil = matcher.__class__.__name__.replace("Matcher", "") + "_LIL"
    #             lil_sheets[payment_name_lil] = df[mask_lil]

    #         if mask_agee.any():
    #             payment_name_agee = matcher.__class__.__name__.replace("Matcher", "") + "_AGEE"
    #             agee_sheets[payment_name_agee] = df[mask_agee]

    #     # Write everything to Excel at once
    #     with pd.ExcelWriter(self.filename, engine='openpyxl', mode='w') as writer:
    #         # Write LIL sheets
    #         for sheet_name, df in lil_sheets.items():
    #             df.to_excel(writer, sheet_name=sheet_name, index=False)
            
    #         # Write AGEE sheets
    #         for sheet_name, df in agee_sheets.items():
    #             df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            
    #         # Concatenate all DataFrames for final processing
    #         df_ordini_payments = pd.concat(all_dfs, ignore_index=True)
    #         df_ordini_payments = df_ordini_payments[df_ordini_payments["Name"].notna()]
    #         df_ordini_payments = df_ordini_payments.groupby("Name", group_keys=False).apply(self.process_check_groups)

    #         # Check for unmatched names
    #         unique_names_payments = df_ordini_payments["Name"].unique()
    #         unmatched_rows = self.df_ordini_iniziale[~self.df_ordini_iniziale["Name"].isin(unique_names_payments)]

    #         # Concatenate payments and unmatched rows
    #         self.df_ordini_all = pd.concat([df_ordini_payments, unmatched_rows], ignore_index=True)
    #         self.df_ordini_all['Total'] = self.df_ordini_all.groupby('Name')['Total'].transform(lambda x: x.where(x.index == x.index[0], np.nan))
    #         self.df_ordini_all = self.df_ordini_all.drop_duplicates(subset=['Name', 'Lineitem name'])

    #         # Checking for inconsistencies in "Total"
    #         inconsistent_totals = self.df_ordini_all.groupby('Name').filter(lambda group: group['Total'].nunique() > 1)
            
    #         if not inconsistent_totals.empty:
    #             print(f"Inconsistent 'Total' values detected for the following 'Names':\n{inconsistent_totals}")
    #         else:
    #             print("All 'Names' have consistent 'Total' values.")

    #         print()

    #         mask_lil = self.df_ordini_all["Brand"] == "LIL Milan"
    #         mask_agee = self.df_ordini_all["Brand"] == "AGEE"

    #         if mask_lil.any():  # Check if there are any True values in the mask
    #             # Write the concatenated data to the main sheet
    #             self.df_ordini_all[mask_lil].to_excel(writer, sheet_name='Ordini LIL', index=False)
            
    #         if mask_agee.any():
    #             self.df_ordini_all[mask_agee].to_excel(writer, sheet_name='Ordini AGEE', index=False)

    #     # Create the summary table in a new sheet
    #     self.create_summary_table()
    #     self.create_daily_summary_table()

    #     # First, create a list of all unique names that have CHECK == "VERO" across all dataframes
    #     all_vero_names = set()  # using a set for unique values
    #     for df in check_dfs:
    #         vero_names = df[df['CHECK'] == "VERO"]['Name'].dropna().unique()
    #         all_vero_names.update(vero_names)

    #     # Now process each dataframe
    #     for df in check_dfs:
    #         mask = (
    #             (~df['Name'].isin(all_vero_names)) | 
    #             (df['Name'].isna()) |
    #             ((df['Name'].isin(all_vero_names)) & (df['CHECK'] == "VERO"))
    #         )
            
    #         # Apply the mask and print
    #         filtered_df = df[mask]

    #         nan_name_df = filtered_df[filtered_df['Name'].isna()]
    #         non_nan_name_df = filtered_df[filtered_df['Name'].notna()].drop_duplicates(subset=['Name', 'Lineitem name'])
    #         filtered_df = pd.concat([nan_name_df, non_nan_name_df], ignore_index=True)
            
    #         if not filtered_df[(filtered_df["CHECK"] != "VERO") & (filtered_df["Brand"] == "LIL Milan")].empty:
    #             display("LIL Milan")
    #             display(filtered_df[(filtered_df["CHECK"] != "VERO") & (filtered_df["Brand"] == "LIL Milan")])

    #         if not filtered_df[(filtered_df["CHECK"] != "VERO") & (filtered_df["Brand"] == "AGEE")].empty:
    #             display("AGEE")
    #             display(filtered_df[(filtered_df["CHECK"] != "VERO") & (filtered_df["Brand"] == "AGEE")])
        
    #     return df_check, self.df_ordini_all, df


    def run_all_matchers(self, mese, anno=2024):
        print("Starting run_all_matchers")
        try:
            qromo_matcher = next((matcher for matcher in self.matchers if isinstance(matcher, QromoMatcher)), None)
            all_dfs = []
            check_dfs = []
            
            print("Processing matchers...")
            for matcher in self.matchers:
                try:
                    print(f"Processing matcher: {type(matcher).__name__}")
                    # Change this comparison
                    if not isinstance(matcher, QromoMatcher):
                        df_check, df_ordini, df = matcher.match()
                    else:
                        df_check, df_ordini, df = matcher.match(mese, anno)
                        
                    print(f"Match successful for {type(matcher).__name__}")
                    all_dfs.append(df_ordini)
                    check_dfs.append(df_check)
                                
                except SkipMatcherException as e:
                    print(f"Skipped matcher: {e}")
                    continue
                except Exception as e:
                    print(f"Error in matcher {type(matcher).__name__}: {str(e)}")
                    raise e

            print("All matchers processed")
            print(f"Number of dfs collected: {len(all_dfs)}")


            # Concatenate all DataFrames for final processing
            df_ordini_payments = pd.concat(all_dfs, ignore_index=True)
            df_ordini_payments = df_ordini_payments[df_ordini_payments["Name"].notna()]
            df_ordini_payments = df_ordini_payments.groupby("Name", group_keys=False).apply(self.process_check_groups)

            # Check for unmatched names using explicit boolean indexing
            unique_names_payments = set(df_ordini_payments["Name"].unique())  # Use a set for faster lookup
            unmatched_mask = ~self.df_ordini_iniziale["Name"].isin(unique_names_payments)
            unmatched_rows = self.df_ordini_iniziale[unmatched_mask]

            # Concatenate payments and unmatched rows
            self.df_ordini_all = pd.concat([df_ordini_payments, unmatched_rows], ignore_index=True)
            self.df_ordini_all['Total'] = self.df_ordini_all.groupby('Name')['Total'].transform(lambda x: x.where(x.index == x.index[0], np.nan))
            self.df_ordini_all = self.df_ordini_all.drop_duplicates(subset=['Name', 'Lineitem name'])

            # Checking for inconsistencies in "Total" using explicit length check
            inconsistent_groups = self.df_ordini_all.groupby('Name').filter(lambda group: group['Total'].nunique() > 1)
            if len(inconsistent_groups) > 0:
                print(f"Inconsistent 'Total' values detected for the following 'Names':\n{inconsistent_groups}")
            else:
                print("All 'Names' have consistent 'Total' values.")

            print()

            # Create a set of all unique names that have CHECK == "VERO" across all dataframes
            all_vero_names = set()
            for df in check_dfs:
                vero_mask = (df['CHECK'] == "VERO") & df['Name'].notna()
                vero_names = set(df[vero_mask]['Name'])
                all_vero_names.update(vero_names)

            non_veri_check = []
            # Now process each dataframe
            for df in check_dfs:
                # Create explicit boolean masks
                not_in_vero = ~df['Name'].isin(all_vero_names)
                is_nan = df['Name'].isna()
                is_vero = (df['Name'].isin(all_vero_names)) & (df['CHECK'] == "VERO")
                
                # Combine masks
                mask = not_in_vero | is_nan | is_vero
                
                # Apply the mask
                filtered_df = df[mask].copy()

                # Split and recombine based on NaN values
                nan_mask = filtered_df['Name'].isna()
                nan_name_df = filtered_df[nan_mask]
                non_nan_name_df = filtered_df[~nan_mask].drop_duplicates(subset=['Name', 'Lineitem name'])
                
                filtered_df = pd.concat([nan_name_df, non_nan_name_df], ignore_index=True)
                non_veri_check.append(filtered_df)

            # Concatenate all filtered dataframes into a single dataframe
            final_df = pd.concat(non_veri_check, ignore_index=True)
            
            # If you want to ensure no duplicates in the final dataframe
            final_df = final_df.drop_duplicates()
            
            
            # Sort by Brand to keep LIL Milan and AGEE records together
            final_df = final_df.sort_values('Brand')

            return final_df
        
        except Exception as e:
            print(f"Error in run_all_matchers: {str(e)}")
            raise e
        
        
            
        
        
        
    # def run_all_matchers(self, mese, anno=2024):
    #     qromo_matcher = next((matcher for matcher in self.matchers if isinstance(matcher, QromoMatcher)), None)
    #     all_dfs = []
    #     check_dfs = []
    #     self.filename = f'Check_{mese:02}_lilmilan.xlsx'
        
    #     # Create dictionaries to store DataFrames for each sheet
    #     lil_sheets = {}
    #     agee_sheets = {}

    #     for matcher in self.matchers:
    #         try:
    #             if matcher != qromo_matcher:
    #                 df_check, df_ordini, df = matcher.match()
    #             else:
    #                 df_check, df_ordini, df = matcher.match(mese, anno)
    #         except SkipMatcherException as e:
    #             print(f"{e}")
    #             continue

    #         # Append to all_dfs and check_dfs
    #         all_dfs.append(df_ordini)
    #         check_dfs.append(df_check)

    #     # Concatenate all DataFrames for final processing
    #     df_ordini_payments = pd.concat(all_dfs, ignore_index=True)
    #     df_ordini_payments = df_ordini_payments[df_ordini_payments["Name"].notna()]
    #     df_ordini_payments = df_ordini_payments.groupby("Name", group_keys=False).apply(self.process_check_groups)

    #     # Check for unmatched names using explicit boolean indexing
    #     unique_names_payments = set(df_ordini_payments["Name"].unique())  # Use a set for faster lookup
    #     unmatched_mask = ~self.df_ordini_iniziale["Name"].isin(unique_names_payments)
    #     unmatched_rows = self.df_ordini_iniziale[unmatched_mask]

    #     # Concatenate payments and unmatched rows
    #     self.df_ordini_all = pd.concat([df_ordini_payments, unmatched_rows], ignore_index=True)
    #     self.df_ordini_all['Total'] = self.df_ordini_all.groupby('Name')['Total'].transform(lambda x: x.where(x.index == x.index[0], np.nan))
    #     self.df_ordini_all = self.df_ordini_all.drop_duplicates(subset=['Name', 'Lineitem name'])

    #     # Checking for inconsistencies in "Total" using explicit length check
    #     inconsistent_groups = self.df_ordini_all.groupby('Name').filter(lambda group: group['Total'].nunique() > 1)
    #     if len(inconsistent_groups) > 0:
    #         print(f"Inconsistent 'Total' values detected for the following 'Names':\n{inconsistent_groups}")
    #     else:
    #         print("All 'Names' have consistent 'Total' values.")

    #     print()

    #     # Create a set of all unique names that have CHECK == "VERO" across all dataframes
    #     all_vero_names = set()
    #     for df in check_dfs:
    #         vero_mask = (df['CHECK'] == "VERO") & df['Name'].notna()
    #         vero_names = set(df[vero_mask]['Name'])
    #         all_vero_names.update(vero_names)

    #     non_veri_check = []
    #     # Now process each dataframe
    #     for df in check_dfs:
    #         # Create explicit boolean masks
    #         not_in_vero = ~df['Name'].isin(all_vero_names)
    #         is_nan = df['Name'].isna()
    #         is_vero = (df['Name'].isin(all_vero_names)) & (df['CHECK'] == "VERO")
            
    #         # Combine masks
    #         mask = not_in_vero | is_nan | is_vero
            
    #         # Apply the mask
    #         filtered_df = df[mask].copy()

    #         # Split and recombine based on NaN values
    #         nan_mask = filtered_df['Name'].isna()
    #         nan_name_df = filtered_df[nan_mask]
    #         non_nan_name_df = filtered_df[~nan_mask].drop_duplicates(subset=['Name', 'Lineitem name'])
            
    #         filtered_df = pd.concat([nan_name_df, non_nan_name_df], ignore_index=True)
    #         non_veri_check.append(filtered_df)

    #     # Concatenate all filtered dataframes into a single dataframe
    #     final_df = pd.concat(non_veri_check, ignore_index=True)
        
    #     # If you want to ensure no duplicates in the final dataframe
    #     final_df = final_df.drop_duplicates()
        
    #     # Sort by Brand to keep LIL Milan and AGEE records together
    #     final_df = final_df.sort_values('Brand')
        
    #     return final_df

    
            # if not filtered_df[(filtered_df["CHECK"] != "VERO") & (filtered_df["Brand"] == "LIL Milan")].empty:
            #     display("LIL Milan")
            #     display(filtered_df[(filtered_df["CHECK"] != "VERO") & (filtered_df["Brand"] == "LIL Milan")])

            # if not filtered_df[(filtered_df["CHECK"] != "VERO") & (filtered_df["Brand"] == "AGEE")].empty:
            #     display("AGEE")
            #     display(filtered_df[(filtered_df["CHECK"] != "VERO") & (filtered_df["Brand"] == "AGEE")])
        




##################### DOPO AGGIUSTAMENTO ############################
            # Create masks
            # mask_lil = df["Brand"] == "LIL Milan"
            # mask_agee = df["Brand"] == "AGEE"

            # # Store filtered DataFrames in dictionaries
            # if mask_lil.any():
            #     payment_name_lil = matcher.__class__.__name__.replace("Matcher", "") + "_LIL"
            #     lil_sheets[payment_name_lil] = df[mask_lil]

            # if mask_agee.any():
            #     payment_name_agee = matcher.__class__.__name__.replace("Matcher", "") + "_AGEE"
            #     agee_sheets[payment_name_agee] = df[mask_agee]

        # # Write everything to Excel at once
        # with pd.ExcelWriter(self.filename, engine='openpyxl', mode='w') as writer:
        #     # Write LIL sheets
        #     for sheet_name, df in lil_sheets.items():
        #         df.to_excel(writer, sheet_name=sheet_name, index=False)
            
        #     # Write AGEE sheets
        #     for sheet_name, df in agee_sheets.items():
        #         df.to_excel(writer, sheet_name=sheet_name, index=False)

        
        # mask_lil = self.df_ordini_all["Brand"] == "LIL Milan"
        # mask_agee = self.df_ordini_all["Brand"] == "AGEE"

        # if mask_lil.any():  # Check if there are any True values in the mask
        #     # Write the concatenated data to the main sheet
        #     self.df_ordini_all[mask_lil].to_excel(writer, sheet_name='Ordini LIL', index=False)
        
        # if mask_agee.any():
        #     self.df_ordini_all[mask_agee].to_excel(writer, sheet_name='Ordini AGEE', index=False)
            
    ############################## FINE AGGIUNTA ################ 
            

        # # Create the summary table in a new sheet
        # self.create_summary_table()
        # self.create_daily_summary_table()

        
        # return df_check, self.df_ordini_all, df

    #check
    def process_check_groups(self, group):
        if "VERO" in group["CHECK"].values:
            return group[group["CHECK"] == "VERO"]
        else:
            return group


    # Function to create stats for each store
    def create_location_stats(self, df, start_row, summary_sheet, store_name):
        # Create groupby stats for this store
        location_stats = df.groupby('Location').agg({'Name': 'nunique',
                                                     'Lineitem quantity': 'sum'
                                                       }).reset_index()
                                
        # Convert to dictionary
        stats_dict = location_stats.set_index('Location').to_dict()
        
        # Get unique locations for this store
        title_of_locations = df["Location"].unique()

        summary_sheet.merge_cells(f'H{start_row-1}:L{start_row-1}') 
        cell = summary_sheet[f'H{start_row-1}']
        cell.value = store_name
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for idx, location_label in enumerate(title_of_locations, start=start_row):
            # Location name
            summary_sheet[f'H{idx}'] = location_label
            
            # SUMIFS formula (adjusted for each store's sheet)
            summary_sheet[f'I{idx}'] = (
                f'=SUMIFS(\'Ordini {store_name}\'!$L:$L, '
                f'\'Ordini {store_name}\'!$BB:$BB, "{location_label}")'
            )
            
            # Stats from dictionary
            unique_orders = stats_dict['Name'].get(location_label, 0)
            items_quantity = stats_dict['Lineitem quantity'].get(location_label, 0)
            
            # Write values
            summary_sheet[f'J{idx}'] = unique_orders
            summary_sheet[f'K{idx}'] = items_quantity
            summary_sheet[f'L{idx}'] = f'=K{idx}/J{idx}'

        
        # Add the "Total" row for locations
        summary_sheet[f'H{idx+1}'] = 'Totale'
        summary_sheet[f'H{idx+1}'].font = Font(bold=True)
        summary_sheet[f'I{idx+1}'] = f'=SUM(I{start_row}:I{idx})'  # Adjust row number based on the last location method
        summary_sheet[f'J{idx+1}'] = f'=SUM(J{start_row}:J{idx})'
        summary_sheet[f'K{idx+1}'] = f'=SUM(K{start_row}:K{idx})'
        summary_sheet[f'L{idx+1}'] = f'=K{idx+1}/J{idx+1}'
        
        # Return the next available row
        return start_row + len(title_of_locations) + 3  # +2 for spacing between tables


    def create_summary_table(self):
        try:
            workbook = load_workbook(self.filename)
        except FileNotFoundError:
            workbook = Workbook()

        # Create or select the sheet for the summary
        summary_sheet = workbook['Totale'] if 'Totale' in workbook.sheetnames else workbook.create_sheet('Totale')

        # Clear previous content in the summary sheet
        for row in summary_sheet.iter_rows(min_row=1, max_col=20, max_row=summary_sheet.max_row):
            for cell in row:
                cell.value = None
                
        # Create bold font style
        bold_font = Font(bold=True)

        # Write headers for payments with bold style
        headers = {
            'A1': 'Payments',
            'B1': 'LIL',
            'C1': 'AGEE',
            'D1': 'CHECK LIL',
            'E1': 'CHECK AGEE'
        }

        for cell_position, value in headers.items():
            cell = summary_sheet[cell_position]
            cell.value = value
            cell.font = bold_font
            
        # Title mapping for totals
        title_of_totals = {
            'Scalapay': 'J',
            'Shopify': 'I',
            'PayPal': 'H',
            'Bonifico': 'H', 
            'Qromo': 'F',
            'Satispay': 'E',
            'Cash': ''  # Placeholder; no column for Cash since it will be a hyphen
        }

        # Write payment methods to the summary sheet
        row = 2  # Start from the second row
        for payment_label, payment_amount in title_of_totals.items():
            summary_sheet[f'A{row}'] = payment_label
            summary_sheet[f'B{row}'] = f'=SUMIFS(\'Ordini LIL\'!$L:$L, \'Ordini LIL\'!$AV:$AV, "*{payment_label}*", \'Ordini LIL\'!$CB:$CB, "LIL Milan")'
            # Add IFERROR to handle missing AGEE sheet
            summary_sheet[f'C{row}'] = f'=IFERROR(SUMIFS(\'Ordini AGEE\'!$L:$L, \'Ordini AGEE\'!$AV:$AV, "*{payment_label}*", \'Ordini AGEE\'!$CB:$CB, "AGEE"), 0)'
            
            if payment_label == "Cash":
                summary_sheet[f'D{row}'] = '-'  # Write a hyphen for Cash
                summary_sheet[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                summary_sheet[f'E{row}'] = '-'  # Write a hyphen for Cash
                summary_sheet[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
            else:
                summary_sheet[f'D{row}'] = f'=IFERROR(SUM(\'{payment_label}_LIL\'!{payment_amount}:{payment_amount}), 0)'
                summary_sheet[f'E{row}'] = f'=IFERROR(SUM(\'{payment_label}_AGEE\'!{payment_amount}:{payment_amount}), 0)'

            row += 1  # Increment the row for the next payment method

        # Add the "Total" row for payments
        summary_sheet[f'A{row}'] = 'Totale'
        summary_sheet[f'A{row}'].font = Font(bold=True)
        summary_sheet[f'B{row}'] = f'=SUM(B2:B{row-1})'  # Adjust row number based on the last payment method
        summary_sheet[f'C{row}'] = f'=SUM(C2:C{row-1})'  # Adjust row number based on the last payment method

        # Leave some columns between the two tables (e.g., start the location table at column E)
        headers = {
            'H1': 'Locations',
            'I1': 'Incasso',
            'J1': 'Ordini',
            'K1': 'Items',
            'L1': 'Oggetti per ordine'
        }

        for cell_position, value in headers.items():
            cell = summary_sheet[cell_position]
            cell.value = value
            cell.font = bold_font

        # Create list of strings to exclude
        exclude_strings = [
            "Luxury Pack",
            "Engraving",
            "E-gift",
            "Repair",
            "Whatever Tote",
            "Piercing Party",
            "LIL Bag"
        ]

        # Create the filter using | (OR) operator between all patterns
        df_ordini_gioielli = self.df_ordini_all[~self.df_ordini_all['Lineitem name'].str.contains('|'.join(exclude_strings), case=False, na=False)]
    
        df_lil = df_ordini_gioielli[df_ordini_gioielli['Brand'] == 'LIL Milan']
        df_agee = df_ordini_gioielli[df_ordini_gioielli['Brand'] == 'AGEE']

        # Create tables for both stores
        start_row = 3  # Starting row for first table
        start_row = self.create_location_stats(df_lil, start_row, summary_sheet, 'LIL')
        if len(df_agee) > 0:
            self.create_location_stats(df_agee, start_row, summary_sheet, 'AGEE')

        # Save the workbook
        workbook.save(self.filename)

    def create_daily_summary_table(self):
        try:
            workbook = load_workbook(self.filename)
        except FileNotFoundError:
            workbook = Workbook()

        # Create or select the sheet for the daily summary
        daily_sheet = workbook['Totale_daily'] if 'Totale_daily' in workbook.sheetnames else workbook.create_sheet('Totale_daily')

        # Clear previous content in the daily sheet
        for row in daily_sheet.iter_rows(min_row=1, max_col=10, max_row=daily_sheet.max_row):
            for cell in row:
                cell.value = None

        # Create bold font style
        bold_font = Font(bold=True)

        # Write headers for payments with bold style
        headers = {
            'A1': 'Giorno',
            'B1': 'Paese',
            'C1': 'Incasso LIL',
            'D1': 'Incasso AGEE',
            'E1': 'Incasso Totale'
        }

        for cell_position, value in headers.items():
            cell = daily_sheet[cell_position]
            cell.value = value
            cell.font = bold_font

        # Extract just the date part (without time) from 'Paid at'
        self.df_ordini_all['Paid at'] = pd.to_datetime(self.df_ordini_all['Paid at'])
        self.df_ordini_all['Data'] = self.df_ordini_all['Paid at'].dt.date

        # First filter for locations
        df_ordini_locations = self.df_ordini_all[self.df_ordini_all["Location"].isin(["Firgun House", "LIL House", "LIL House London"])]

        # Split by store
        df_ordini_locations_lil = df_ordini_locations[df_ordini_locations["Brand"] == "LIL Milan"]
        df_ordini_locations_agee = df_ordini_locations[df_ordini_locations["Brand"] == "AGEE"]

        # Calculate totals for LIL
        daily_country_totals_lil = df_ordini_locations_lil.groupby(['Data', 'Shipping Country'])['Total'].sum().reset_index()
        daily_country_totals_lil = daily_country_totals_lil.sort_values(['Data', 'Shipping Country'])
        daily_country_totals_lil = daily_country_totals_lil.rename(columns={'Shipping Country': 'Country', 'Total': 'Total_LIL'})

        # Calculate totals for AGEE
        daily_country_totals_agee = df_ordini_locations_agee.groupby(['Data', 'Shipping Country'])['Total'].sum().reset_index()
        daily_country_totals_agee = daily_country_totals_agee.sort_values(['Data', 'Shipping Country'])
        daily_country_totals_agee = daily_country_totals_agee.rename(columns={'Shipping Country': 'Country', 'Total': 'Total_AGEE'})

        daily_country_totals = pd.merge(daily_country_totals_lil, daily_country_totals_agee, on=["Data", "Country"], how="outer")
        daily_country_totals = daily_country_totals.fillna(0)

        # Write to Excel
        for idx, row in enumerate(daily_country_totals.itertuples(), start=2):
            daily_sheet[f'A{idx}'] = row.Data
            daily_sheet[f'B{idx}'] = row.Country
            daily_sheet[f'C{idx}'] = row.Total_LIL      # LIL total
            daily_sheet[f'D{idx}'] = row.Total_AGEE     # AGEE total
            daily_sheet[f'E{idx}'] = f'=C{idx}+D{idx}'

        # Add totals row at the bottom
        last_row = idx + 2
        daily_sheet[f'C{last_row}'] = f'=SUM(C2:C{idx})'
        daily_sheet[f'D{last_row}'] = f'=SUM(D2:D{idx})'
        daily_sheet[f'E{last_row}'] = f'=SUM(E2:E{idx})'
    
        # Save the workbook
        workbook.save(self.filename)
