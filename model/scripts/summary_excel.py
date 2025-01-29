#CLASSE PER GENERARE L'EXCEL:

import pandas as pd
import numpy as np
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


class OrderSummary:
    def __init__(self, df_ordini_all, pagamenti, filename):
        self.df_ordini_all = df_ordini_all
        self.pagamenti = pagamenti
        self.filename = filename

    def process_group(self, group):
        # Check if all non-NaN 'Total' values in the group are the same
        unique_totals = group['Total'].dropna().unique()
        
        if len(unique_totals) == 1:
            # If all 'Total' values are the same, keep only the first non-NaN occurrence
            group.loc[group['Total'].notna().cumsum() > 1, 'Total'] = pd.NA
        # If values are different, keep them as they are (no changes)
        return group
    
    # # Function to process data for a specific brand
    # def process_location_df(self, df, brand, exclude_strings):
    #     df_filtered = df[(df['Brand'] == brand) 
    #                     & (df['CHECK'] != 'ESCLUSO') 
    #                     & ((df['Total'] != 0) | ((df['Total'] == 0) & (df['Payment Method'] == "Gift Card")))]
        
    #     # Optimize 'Lineitem quantity gioiello' calculation
    #     df_filtered['Lineitem quantity gioiello'] = np.where(
    #         pd.isna(df_filtered['Lineitem name']) | (~df_filtered['Lineitem name'].str.contains('|'.join(exclude_strings), na=False)),
    #         df_filtered['Lineitem quantity'], #if true
    #         0 #if false
    #     )
        
    #     # Calculate 'Lineitem quantity gioiello per name'
    #     df_filtered['Lineitem quantity gioiello per name'] = df_filtered.groupby('Name')['Lineitem quantity gioiello'].transform('sum')
        
    #     # Group by 'Location'
    #     group_location = df_filtered.groupby('Location').agg({
    #         'Name': 'nunique',
    #         'Lineitem quantity': 'sum',
    #         'Lineitem quantity gioiello': 'sum'
    #     }).reset_index()
        
    #     # Filter and group names with non-zero gioielli
    #     gioielli = df_filtered[df_filtered['Lineitem quantity gioiello per name'] > 0].drop_duplicates('Name')
    #     gioielli_count = gioielli.groupby('Location')['Name'].nunique()
    #     gioielli_count.name = 'Name solo gioielli'
        
    #     # Merge the results
    #     result = pd.merge(group_location, gioielli_count, on='Location', how='left').fillna(0)
    #     return result

    def check_names_pagamenti(self, df_pagamenti):

        # df_pagamenti = df_pagamenti.sort_values(by = "Data")
        grouped_pagamenti = (df_pagamenti.groupby("Name", as_index=False).agg({"Importo Pagato": "sum",  # Sum over 'Importo Pagato'
                                                                                "Numero Pagamento": "first",   # Take the first value of 'description'
                                                                                "Metodo": "first"
                                                                            }))

        # Iterate over each unique Name in df_ordini
        for name in grouped_pagamenti["Name"].unique():
            num = grouped_pagamenti[grouped_pagamenti["Name"] == name]["Numero Pagamento"].values[0]
            pag = grouped_pagamenti[grouped_pagamenti["Name"] == name]["Metodo"].values[0]
            if pag == "Qromo" or (pag == "Satispay" and num == 0):
                if name in self.df_ordini_all["Name"].values:
                    total_importo = grouped_pagamenti.loc[grouped_pagamenti["Name"] == name, "Importo Pagato"].values[0]
                    total_ordini = self.df_ordini_all.loc[self.df_ordini_all["Name"] == name, "Total"].values[0]

                    # If total_importo == total_ordini
                    if total_importo == total_ordini:
                        df_pagamenti.loc[df_pagamenti["Name"] == name, "CHECK"] = "VERO"

                    # If totals do not match, check individual rows
                    else:
                        name_rows = df_pagamenti[df_pagamenti["Name"] == name]

                        for _, row in name_rows.iterrows():
                            if row["Importo Pagato"] == total_ordini:
                                # Mark the matched row as CHECK == VERO
                                df_pagamenti.loc[(df_pagamenti["Name"] == name) & (df_pagamenti["Importo Pagato"] == total_ordini), "CHECK"] = "VERO"
                            else:
                                df_pagamenti.loc[(df_pagamenti["Name"] == name) & (df_pagamenti["Importo Pagato"] != total_ordini), ["CHECK", "Name"]] = ["NON TROVATO", None]

                else:
                    df_pagamenti.loc[df_pagamenti["Name"] == name, "CHECK"] = "NON TROVATO"


        for numero, nome in st.session_state.pagamenti_da_aggiungere_lil.items():
            df_pagamenti.loc[(df_pagamenti["Numero Pagamento"] == numero), ["CHECK", "Name"]] = ["VERO", nome]  

        for numero, nome in st.session_state.pagamenti_da_aggiungere_agee.items():
            df_pagamenti.loc[(df_pagamenti["Numero Pagamento"] == numero), ["CHECK", "Name"]] = ["VERO", nome]            

        return df_pagamenti


    def create_files(self):

        # Define exclude strings
        exclude_strings = ["Luxury Pack", "Engraving", "E-Gift", "Repair", "Whatever Tote", "Piercing Party", "LIL Bag"]

        try:
            sheet_order = ['Totale',           # First summary sheet
                            'Totale_daily',     # Daily summary sheet
                            'Ordini LIL',       # Orders sheets
                            'Ordini AGEE',
                            'Bonifico_LIL',
                            'Cash_LIL', 
                            'PayPal_LIL', 
                            'Qromo_LIL', 
                            'Satispay_LIL', 
                            'Scalapay_LIL', 
                            'Shopify_LIL', 
                            'Bonifico_AGEE', 'PayPal_AGEE', 'Qromo_AGEE',
                            'Satispay_AGEE', 'Scalapay_AGEE', 'Shopify_AGEE'
                        ]

            # Create a new workbook first
            wb = Workbook()
            wb.save(self.filename)

            # columns_state = ColumnsState.get_instance()
            paid_at_pos = st.session_state.df_columns.get_loc('Paid at')
            created_at_pos = st.session_state.df_columns.get_loc('Created at')
            df_columns = list(st.session_state.df_columns)
            df_columns.insert(created_at_pos + 1, 'Lineitem gioiello')
            df_columns.insert(paid_at_pos + 1, 'Data Giorno')
            
            pagamenti_columns = st.session_state.pagamenti_columns
            for metodo, columns in pagamenti_columns.items():
                pagamenti_columns[metodo] = list(columns)
                # pagamenti_columns[metodo].append("CHECK")

            lista_ordini = st.session_state.pagamenti["Name"].unique().tolist()
            for nome in st.session_state.pagamenti_da_aggiungere_lil.values():
                if nome not in lista_ordini:
                    lista_ordini.append(nome)
            for nome in st.session_state.pagamenti_da_aggiungere_agee.values():
                if nome not in lista_ordini:
                    lista_ordini.append(nome)
                        
            # Apply the function to each 'Name' group
            self.df_ordini_all = self.df_ordini_all.groupby('Name', group_keys=False).apply(self.process_group)

            # First write the basic DataFrames
            with pd.ExcelWriter(self.filename, engine='openpyxl', mode='a') as writer:
                mask_lil_o = self.df_ordini_all["Brand"] == "Ordini LIL"
                mask_agee_o = self.df_ordini_all["Brand"] == "Ordini AGEE"

                # Write order sheets first (these are needed for the summary tables)
                if mask_lil_o.any():
                    lil = self.df_ordini_all[mask_lil_o].copy()
                     
                    # Create "Lineitem gioiello" column based on condition
                    created_at_index = lil.columns.get_loc("Created at")
                    lil.insert(created_at_index + 1, "Lineitem gioiello", lil["Lineitem name"].apply(
                        lambda x: False if any(exclude.lower() in x.lower() for exclude in exclude_strings) else True
                    ))

                    # Create "Data Giorno" column
                    paid_at_index = lil.columns.get_loc("Paid at")
                    lil.insert(paid_at_index + 1, "Data Giorno", lil["Paid at"].apply(self.reformat_date))
                    
                    # Filter columns to keep only the desired ones
                    lil = lil[df_columns]
                    
                    # Write to Excel
                    lil.to_excel(writer, sheet_name='Ordini LIL', index=False)

                if mask_agee_o.any():
                    agee = self.df_ordini_all[mask_agee_o].copy()

                    created_at_index = agee.columns.get_loc("Created at")
                    agee.insert(created_at_index + 1, "Lineitem gioiello", agee["Lineitem name"].apply(
                        lambda x: False if any(exclude.lower() in x.lower() for exclude in exclude_strings) else True
                    ))
                
                    paid_at_index = agee.columns.get_loc("Paid at")
                    agee.insert(paid_at_index + 1, "Data Giorno", agee["Paid at"].apply(self.reformat_date))

                    agee = agee[df_columns]

                    # Write to Excel
                    agee.to_excel(writer, sheet_name='Ordini AGEE', index=False)

                # Write payment sheets
                mask_lil_p = self.pagamenti["Brand"] == "Ordini LIL"
                mask_agee_p = self.pagamenti["Brand"] == "Ordini AGEE"

                if mask_lil_p.any():
                    for p in self.pagamenti["Metodo"].unique():
                        payment_name_lil = p.split()[0] + "_LIL"

                        if p == "Cash": 
                            filtered_df_lil = self.pagamenti[self.pagamenti["Metodo"] == p]
                        else:
                            filtered_df_lil = self.pagamenti[mask_lil_p & (self.pagamenti["Metodo"] == p)]

                        if not filtered_df_lil.empty:
                            matching_columns = next((cols for key, cols in pagamenti_columns.items() 
                                                    if key == p), None)

                            if len(matching_columns) > 0:
                                if p == "Cash":
                                    filtered_df_lil = filtered_df_lil[matching_columns]
                                
                                else:
                
                                    if p == "PayPal Express Checkout" or p == "Qromo" or p == "Bonifico":
                                        desired_columns = matching_columns + ["Numero Pagamento", "Name", "Importo Pagato", "Metodo", "CHECK"]
                                        filtered_df_lil = filtered_df_lil[desired_columns]
                                        filtered_df_lil = self.check_names_pagamenti(filtered_df_lil)

                                        filtered_df_lil = filtered_df_lil.drop(["Numero Pagamento", "Importo Pagato", "Metodo"], axis = 1)
                                    else:
                                        desired_columns = matching_columns + ["Data", "Numero Pagamento", "Name", "Importo Pagato", "Metodo", "CHECK"]
                                        filtered_df_lil = filtered_df_lil[desired_columns]
                                        filtered_df_lil = self.check_names_pagamenti(filtered_df_lil)

                                        filtered_df_lil = filtered_df_lil.drop(["Data", "Numero Pagamento", "Importo Pagato", "Metodo"], axis = 1)

                                    blank_col_index = len(matching_columns)
                                    filtered_df_lil.insert(blank_col_index, "", "")

                            filtered_df_lil.to_excel(writer, sheet_name=payment_name_lil, index=False)

                if mask_agee_p.any():
                    for p in self.pagamenti["Metodo"].unique():
                        payment_name_agee = p.split()[0] + "_AGEE"
                        filtered_df_agee = self.pagamenti[mask_agee_p & (self.pagamenti["Metodo"] == p)]
                        
                        if not filtered_df_agee.empty:
                            matching_columns = next((cols for key, cols in pagamenti_columns.items() 
                                                if key == p), None)

                            # Filter columns if match found
                            if len(matching_columns) > 0:
                                desired_columns = matching_columns + ["Numero Pagamento", "Name", "Importo Pagato", "Metodo", "CHECK"]
                                filtered_df_agee = filtered_df_agee[desired_columns]

                                filtered_df_agee = self.check_names_pagamenti(filtered_df_agee)
                                filtered_df_agee = filtered_df_agee.drop(["Numero Pagamento", "Importo Pagato", "Metodo"], axis = 1)

                                blank_col_index = len(matching_columns)
                                filtered_df_agee.insert(blank_col_index, "", "")
                    
                            filtered_df_agee.to_excel(writer, sheet_name=payment_name_agee, index=False)

            # Highlight rows where CHECK != 'VERO'
            wb = load_workbook(self.filename)

            for sheetname in wb.sheetnames:

                # Check the "CHECK" column and highlight rows where the value is not "VERO"
                check_col = None

                name_col = None
                total_col = None
                payment_method_col = None
                
                sheet = wb[sheetname]

                for col in sheet.iter_cols(1, sheet.max_column):
                    if col[0].value == "CHECK":
                        check_col = col[0].column
                        break

                if check_col:
                    for row in sheet.iter_rows(2, sheet.max_row):
                        if row[check_col - 1].value != "VERO":
                            for cell in row:
                                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                if sheetname.startswith("Ordini"):
                    for col in sheet.iter_cols(1, sheet.max_column):
                        if col[0].value == "Name":
                            name_col = col[0].column
                        elif col[0].value == "Total":
                            total_col = col[0].column
                        elif col[0].value == "Payment Method":
                            payment_method_col = col[0].column

                        # Break early if all required columns are found
                        if name_col and total_col and payment_method_col:
                            break

                    if name_col and total_col and payment_method_col:
                        # Iterate through all rows starting from the second row (to skip headers)
                        for row in sheet.iter_rows(2, sheet.max_row):
                            name_value = row[name_col - 1].value
                            total_value = row[total_col - 1].value
                            payment_method_value = row[payment_method_col - 1].value

                            # Check if "Name" is not in lista_ordini, "Total" is not NaN or 0, and "Payment Method" is not "Cash"
                            if name_value not in lista_ordini and (total_value is not None and total_value != 0 and payment_method_value != "Cash"):
                                # Highlight the row with yellow
                                for cell in row:
                                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Save the workbook after all changes
            wb.save(self.filename)

            # Now create the summary tables after all required sheets exist
            self.create_summary_table()  # This creates 'Totale' sheet
            self.create_daily_summary_table()  # This creates 'Totale_daily' sheet

            # Finally reorder sheets
            workbook = load_workbook(self.filename)

            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

            self.reorder_sheets(workbook, sheet_order)
            workbook.save(self.filename)
                
            return self.filename

        except Exception as e:
            print(f"Error creating Excel file: {str(e)}")
            raise


    def reorder_sheets(self, workbook, sheet_order):
        """Helper function to reorder sheets in the workbook"""
        try:
            # Get the sheets that exist in the workbook
            existing_sheets = set(workbook.sheetnames)
            
            # Filter sheet_order to only include sheets that exist
            final_order = [sheet for sheet in sheet_order if sheet in existing_sheets]
            
            # Add any existing sheets that weren't in the order list at the end
            final_order.extend(sheet for sheet in existing_sheets if sheet not in final_order)
            
            # Reorder the sheets
            for i, sheet_name in enumerate(final_order):
                current_index = workbook.index(workbook[sheet_name])
                if current_index != i:  # Only move if needed
                    workbook.move_sheet(sheet_name, offset=i-current_index)
                    
        except Exception as e:
            print(f"Warning: Error while reordering sheets: {str(e)}")
            print("Continuing with original sheet order...")

    # Function to check and reformat
    def reformat_date(self, date_str):
        if pd.notna(date_str):
            date_str = date_str.strip().replace("/", "-")[:10]
        # Check if the string starts with "2024"
            if not date_str.startswith("2024"):
                # Split and rearrange as "YYYY-MM-DD"
                return "-".join(date_str.split("-")[::-1])
        return date_str  # Return as-is if already starts with "2024"

 
    # Method to create stats for each store location
    def create_location_stats(self, wb, start_row, summary_sheet, store_name):

        # df['Lineitem quantity'] = df['Lineitem quantity'].astype(int)
        # df['Lineitem quantity gioiello'] = df['Lineitem quantity gioiello'].astype(int)
        # # # location_stats = df.groupby('Location').agg({'Name': 'nunique',
        # # #                                              'Lineitem quantity': 'sum'}).reset_index()
        # # # stats_dict = location_stats.set_index('Location').to_dict()

        # # # Get unique locations in the desired order
        if store_name == "LIL":
            title_of_locations = ["Firgun House", "LIL House", "LIL House London"]
        else:
            # Check if 'Ordini AGEE' sheet exists
            if 'Ordini AGEE' in wb.sheetnames:
                title_of_locations = ["Firgun House", "LIL House"]
            else:
                print("Sheet 'Ordini AGEE' does not exist, breaking the process.")
                return  # This will stop the code execution

        # Get unique values from 'Ordini LIL' sheet in column AX
        ordini_sheet = wb['Ordini ' + store_name]  # Load the 'Ordini LIL' sheet
        locations_sheet = [cell.value for cell in ordini_sheet['BD'] if cell.value is not None]
        unique_locations_sheet = set(locations_sheet)

        locations = [location for location in unique_locations_sheet if location in title_of_locations]
        locations.sort()

        # Prepare the summary sheet
        summary_sheet.merge_cells(f'L{start_row-1}:R{start_row-1}') 
        cell = summary_sheet[f'L{start_row-1}']
        cell.value = store_name
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # # Group the DataFrame by Location to ensure data alignment
        # grouped_df = df.groupby("Location")

        # # Iterate over the unique locations
        # for idx, location_label in enumerate(title_of_locations, start=start_row):
        #     # Write the location name
        #     

        #     # Add formula for total quantity based on the location
        

        #     # Retrieve corresponding data for the location
        #     location_data = grouped_df.get_group(location_label) if location_label in grouped_df.groups else None
        #     if location_data is not None:
                # Fill in the data
        last_row = None
        for idx, location_label in enumerate(locations, start=start_row):
            summary_sheet[f'L{idx}'] = location_label
            summary_sheet[f'M{idx}'] = f'=SUMIFS(\'Ordini {store_name}\'!$M:$M, \'Ordini {store_name}\'!$BD:$BD, $L{idx})'
            summary_sheet[f'N{idx}'] = (f'=SUMPRODUCT(((\'Ordini {store_name}\'!$M:$M > 0) + ((\'Ordini {store_name}\'!$M:$M = 0) * (\'Ordini {store_name}\'!$AX:$AX = "Gift Card"))) *'
                                        f'(\'Ordini {store_name}\'!$BD:$BD = $L{idx}))')
            summary_sheet[f'O{idx}'] = (f'=SUMPRODUCT(((\'Ordini {store_name}\'!$M$2:$M$20000 > 0) +' 
                                            f'(\'Ordini {store_name}\'!$M$2:$M$20000 = "") +'
                                            f'((\'Ordini {store_name}\'!$M$2:$M$20000 = 0) * (\'Ordini {store_name}\'!$AX$2:$AX$20000 = "Gift Card")) > 0) *'
                                        f'(\'Ordini {store_name}\'!$BD$2:$BD$20000 = $L{idx}) *'
                                        f'(\'Ordini {store_name}\'!$S$2:$S$20000 > 0) *'
                                        f'(\'Ordini {store_name}\'!$S$2:$S$20000) *'
                                        f'(ISERROR(SEARCH("100", \'Ordini {store_name}\'!$N$2:$N$20000))))')
            summary_sheet[f'P{idx}'] = (f'=SUMPRODUCT(((\'Ordini {store_name}\'!$M$2:$M$20000 > 0) +'
                                            f'(\'Ordini {store_name}\'!$M$2:$M$20000 = "") +'
                                            f'((\'Ordini {store_name}\'!$M$2:$M$20000 = 0) * (\'Ordini {store_name}\'!$AX$2:$AX$20000 = "Gift Card")) > 0) *'
                                        f'(\'Ordini {store_name}\'!$R$2:$R20000 = TRUE) *'
                                        f'(\'Ordini {store_name}\'!$BD$2:$BD$20000 = $L{idx}) *'
                                        f'(\'Ordini {store_name}\'!$S$2:$S$20000 > 0) *'
                                        f'(\'Ordini {store_name}\'!$S$2:$S$20000) *'
                                        f'(ISERROR(SEARCH("100", \'Ordini {store_name}\'!$N$2:$N$20000))))')
            summary_sheet[f'Q{idx}'] = f'=O{idx}/N{idx}'
            summary_sheet[f'Q{idx}'].number_format = '0.00'
            summary_sheet[f'R{idx}'] = f'=P{idx}/N{idx}'
            summary_sheet[f'R{idx}'].number_format = '0.00'


            last_row = idx

        if last_row is not None:
            summary_sheet[f'L{last_row+1}'] = 'Totale'
            summary_sheet[f'L{last_row+1}'].font = Font(bold=True)
            summary_sheet[f'M{last_row+1}'] = f'=SUM(M{start_row}:M{last_row})'
            summary_sheet[f'M{last_row+1}'].font = Font(bold=True)
            summary_sheet[f'N{last_row+1}'] = f'=SUM(N{start_row}:N{last_row})'
            summary_sheet[f'O{last_row+1}'] = f'=SUM(O{start_row}:O{last_row})'
            summary_sheet[f'P{last_row+1}'] = f'=SUM(P{start_row}:P{last_row})'
            summary_sheet[f'Q{last_row+1}'] = f'=O{last_row+1}/N{last_row+1}'
            summary_sheet[f'Q{last_row+1}'].number_format = '0.00'
            summary_sheet[f'R{last_row+1}'] = f'=P{last_row+1}/N{last_row+1}'
            summary_sheet[f'R{last_row+1}'].number_format = '0.00'


        return start_row + len(locations) + 3

    # Method to create a summary table in Excel
    def create_summary_table(self):
        try:
            workbook = load_workbook(self.filename)
        except FileNotFoundError:
            workbook = Workbook()

        summary_sheet = workbook['Totale'] if 'Totale' in workbook.sheetnames else workbook.create_sheet('Totale')

        for row in summary_sheet.iter_rows(min_row=1, max_col=20, max_row=summary_sheet.max_row):
            for cell in row:
                cell.value = None

        bold_font = Font(bold=True)
        headers = {
            'A1': 'Payments', 'B1': '', 'C1': 'LIL',  'D1': 'CHECK LIL', 'E1': 'DIFF LIL', 'F1': '', 'G1': 'AGEE','H1': 'CHECK AGEE', 'I1': 'DIFF AGEE'}
        for cell_position, value in headers.items():
            cell = summary_sheet[cell_position]
            cell.value = value
            cell.font = bold_font

        title_of_totals = {
            'Scalapay': 'J', 'Shopify': 'I', 'PayPal': 'H', 'Bonifico': 'H',
            'Qromo': 'F', 'Satispay': 'E', 'Cash': 'H'
        }

        row = 2
        for payment_label, payment_amount in title_of_totals.items():
            summary_sheet[f'A{row}'] = payment_label
            summary_sheet[f'C{row}'] = f'=SUMIFS(\'Ordini LIL\'!$M:$M, \'Ordini LIL\'!$AX:$AX, "*{payment_label}*")'
            summary_sheet[f'G{row}'] = f'=IFERROR(SUMIFS(\'Ordini AGEE\'!$M:$M, \'Ordini AGEE\'!$AX:$AX, "*{payment_label}*"), 0)'
            
            if payment_label == "Qromo":
                summary_sheet[f'D{row}'] = f'=IFERROR(SUM(\'{payment_label}_LIL\'!C:C) - SUM(\'{payment_label}_LIL\'!D:D), 0)'
                summary_sheet[f'H{row}'] = f'=IFERROR(SUM(\'{payment_label}_AGEE\'!C:C) - SUM(\'{payment_label}_AGEE\'!D:D), 0)'
                summary_sheet[f'E{row}'] = f'=D{row}-C{row}'
                summary_sheet[f'I{row}'] = f'=H{row}-G{row}'
            
            else:
                summary_sheet[f'D{row}'] = f'=IFERROR(SUM(\'{payment_label}_LIL\'!{payment_amount}:{payment_amount}), 0)'
                summary_sheet[f'H{row}'] = f'=IFERROR(SUM(\'{payment_label}_AGEE\'!{payment_amount}:{payment_amount}), 0)'
                summary_sheet[f'E{row}'] = f'=D{row}-C{row}'
                summary_sheet[f'I{row}'] = f'=H{row}-G{row}'

            row += 1

        summary_sheet[f'A{row}'] = 'Totale'
        summary_sheet[f'A{row}'].font = Font(bold=True)
        summary_sheet[f'C{row}'] = f'=SUM(C2:C{row-1})'
        summary_sheet[f'C{row}'].font = Font(bold=True)
        summary_sheet[f'G{row}'] = f'=SUM(G2:G{row-1})'
        summary_sheet[f'G{row}'].font = Font(bold=True)

        headers = {'L1': 'Locations', 'M1': 'Incasso', 'N1': 'Ordini', 'O1': 'Items', 'P1': 'Gioielli', 'Q1': 'Items per Ordine', 'R1': 'Gioielli per Ordine'}
        for cell_position, value in headers.items():
            cell = summary_sheet[cell_position]
            cell.value = value
            cell.font = bold_font

        # # Apply to both brands
        # exclude_strings = ["Luxury Pack", "Engraving", "E-Gift", "Repair", "Whatever Tote", "Piercing Party", "LIL Bag"]

        # # Sort and fill Total
        # df_ordini_fill = self.df_ordini_all.sort_values(by=['Name', "Total"])
        # df_ordini_fill["Total"] = df_ordini_fill.groupby('Name')["Total"].ffill()

        # # Process LIL and AGEE data
        # df_lil = self.process_location_df(df_ordini_fill, \'Ordini {store_name}\', exclude_strings)
        # df_agee = self.process_location_df(df_ordini_fill, 'Ordini AGEE', exclude_strings)

        start_row = 3
        start_row = self.create_location_stats(workbook, start_row, summary_sheet, 'LIL')
        self.create_location_stats(workbook, start_row, summary_sheet, 'AGEE')

        workbook.save(self.filename)

    # Method to create a daily summary table
    def create_daily_summary_table(self):
        try:
            workbook = load_workbook(self.filename)
        except FileNotFoundError:
            workbook = Workbook()

        daily_sheet = workbook['Totale_daily'] if 'Totale_daily' in workbook.sheetnames else workbook.create_sheet('Totale_daily')

        for row in daily_sheet.iter_rows(min_row=1, max_col=20, max_row=daily_sheet.max_row):
            for cell in row:
                cell.value = None

        bold_font = Font(bold=True)
        # Set static headers for Giorno and Totale
        headers = {'A1': 'Giorno', 'B1': 'Refunded Amount', 'C1':	'Total', 'D1':'', 'E1':	'Totale Effettivo'}

        # Filter the "Ordini LIL" sheet data to get unique Shipping Country values with CHECK != 'ESCLUSO'
        ordini_lil_df = self.df_ordini_all[(self.df_ordini_all['Brand'] == 'Ordini LIL') 
                                           & (self.df_ordini_all['CHECK'] != 'ESCLUSO') 
                                           & (self.df_ordini_all["Location"].isin(["Firgun House", "LIL House", "LIL House London"]))]

        paid_at_index = ordini_lil_df.columns.get_loc("Paid at")
        ordini_lil_df.insert(paid_at_index + 1, "Data Giorno", ordini_lil_df["Paid at"].apply(self.reformat_date))

        unique_countries = ordini_lil_df['Shipping Country'].unique()
        ue_countries  = ['AT', 'BE', 'BG', 'CY', 'HR', 'DK', 'EE', 'FI', 'FR', 'DE', 'EL', 'IE', 'IT', 'LV', 'LT', 'LU', 'MT', 'NL', 'PL', 'PT', 'CZ', 'SK', 'RO', 'SI', 'ES', 'SE', 'HU']

        art8_countries = [code for code in unique_countries if code not in ue_countries]

        # Add static headers
        for cell_position, value in headers.items():
            cell = daily_sheet[cell_position]
            cell.value = value
            cell.font = bold_font

        # Add headers for each unique Shipping Country starting from column 'D'
        start_col = 7  # Column index for 'G'

        #paesi ue 
        i = 0
        for country in unique_countries:
            if country in ue_countries:
                col_letter = chr(64 + start_col + i)  # Convert to Excel column letter
                cell_position = f'{col_letter}1'
                daily_sheet[cell_position] = country
                daily_sheet[cell_position].font = bold_font
                i += 1

        #paesi extra ue tutti messi insieme
        col_letter_art8 = chr(64 + start_col + i)  # Convert to Excel column letter
        print("art8", col_letter_art8)
        cell_position = f'{col_letter_art8}1'
        daily_sheet[cell_position] = "Art.8"
        daily_sheet[cell_position].font = bold_font

        # Get unique dates from the "Giorno" column
        unique_dates = ordini_lil_df['Data Giorno'].dropna().unique()
        unique_dates.sort()  # Sort dates if needed

        # Start filling the first column (A) with these unique dates, beginning at row 2
        for row, giorno in enumerate(unique_dates, start=2):
            daily_sheet[f'A{row}'] = giorno
            daily_sheet[f'B{row}'] = f"=SUMIFS(\'Ordini LIL\'!$AZ:$AZ, \'Ordini LIL\'!$E:$E, A{row})"
            daily_sheet[f'C{row}'] = f"=SUMIFS(\'Ordini LIL\'!$M:$M, \'Ordini LIL\'!$E:$E, A{row}) + SUMIFS(\'Ordini LIL\'!$AZ:$AZ, \'Ordini LIL\'!$E:$E, A{row})"
            daily_sheet[f'E{row}'] = f"=SUMIFS(\'Ordini LIL\'!$M:$M, \'Ordini LIL\'!$E:$E, A{row})"

            #eu contries
            i = 0
            for country in unique_countries:
                if country in ue_countries:
                    col_letter = chr(64 + start_col + i)  # Convert to Excel column letter
                    print(country)
                    daily_sheet[f'{col_letter}{row}'] = f"=SUMIFS(\'Ordini LIL\'!$M:$M, \'Ordini LIL\'!$E:$E, $A{row}, \'Ordini LIL\'!$AS:$AS, {col_letter}$1)" 
                    print(f"=SUMIFS(\'Ordini LIL\'!$M:$M, \'Ordini LIL\'!$E:$E, $A{row}, \'Ordini LIL\'!$AS:$AS, {col_letter}$1)" )
                    i += 1
        
            # ART8 countries
            final_formula = " + ".join([f'SUMIFS(\'Ordini LIL\'!$M:$M, \'Ordini LIL\'!$E:$E, $A{row}, \'Ordini LIL\'!$AS:$AS, \"{country}\")' for country in art8_countries])
            daily_sheet[f'{col_letter_art8}{row}'] = f"={final_formula}"
            print(f"={final_formula}")

            idx = row


        daily_sheet[f'A{idx+2}'] = "Totale"
        daily_sheet[f'A{idx+2}'].font = Font(bold=True)
        daily_sheet[f'B{idx+2}'] = f"=SUM(B2:B{idx})"
        daily_sheet[f'C{idx+2}'] = f"=SUM(C2:C{idx})"
        daily_sheet[f'E{idx+2}'] = f"=SUM(E2:E{idx})"
        daily_sheet[f'E{idx+2}'].font = Font(bold=True)

        #eu countries
        i = 0
        for country in unique_countries:
            if country in ue_countries:
                col_letter = chr(64 + start_col + i)  # Convert to Excel column letter
                daily_sheet[f'{col_letter}{idx+2}'] = f"=SUM({col_letter}2:{col_letter}{idx})"
                i += 1

        #art8
        daily_sheet[f'{col_letter_art8}{idx+2}'] = f"=SUM({col_letter_art8}2:{col_letter_art8}{idx})"

        col_index_art8 = ord(col_letter_art8) - 64

        # Add the final SUM formula for the total
        total_col_letter = chr(64 + col_index_art8 + 2)  # Next column after the last one
        daily_sheet[f'{total_col_letter}{idx+2}'] = f"=SUM(G{idx+2}:{col_letter_art8}{idx+2})"
        daily_sheet[f'{total_col_letter}{idx+2}'].font = Font(bold=True)

        workbook.save(self.filename)



