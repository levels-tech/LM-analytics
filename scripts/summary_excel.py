#CLASSE PER GENERARE L'EXCEL:

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

from utils.columns_state import ColumnsState


class OrderSummary:
    def __init__(self, df_ordini_all, pagamenti, filename):
        self.df_ordini_all = df_ordini_all
        self.pagamenti = pagamenti
        self.filename = filename

    #PROVAAAAAAAA
    def process_group(self, group):
        # Check if all non-NaN 'Total' values in the group are the same
        unique_totals = group['Total'].dropna().unique()
        
        if len(unique_totals) == 1:
            # If all 'Total' values are the same, keep only the first non-NaN occurrence
            group.loc[group['Total'].notna().cumsum() > 1, 'Total'] = pd.NA
        # If values are different, keep them as they are (no changes)
        return group


    def create_files(self):
        try:
            sheet_order = ['Totale',           # First summary sheet
                            'Totale_daily',     # Daily summary sheet
                            'Ordini LIL',       # Orders sheets
                            'Ordini AGEE',
                            'Bonifico_LIL', 
                            'PayPal_LIL', 
                            'Qromo_LIL', 
                            'Satispay_LIL', 
                            'Scalapay_LIL', 
                            'Shopify_LIL', 
                            'Bonifico_AGEE', 'PayPal_AGEE', 'Qromo_AGEE',
                            'Satispay_AGEE', 'Scalapay_AGEE', 'Shopify_AGEE'
                        ]

            # Create a new workbook first
            wb = Workbook()
            wb.save(self.filename)

            columns_state = ColumnsState.get_instance()
            df_columns = columns_state.df_columns
            pagamenti_columns = columns_state.pagamenti_columns

            # Apply the function to each 'Name' group
            self.df_ordini_all = self.df_ordini_all.groupby('Name', group_keys=False).apply(self.process_group)

            # First write the basic DataFrames
            with pd.ExcelWriter(self.filename, engine='openpyxl', mode='a') as writer:
                mask_lil_o = self.df_ordini_all["Brand"] == "LIL Milan"
                mask_agee_o = self.df_ordini_all["Brand"] == "AGEE"


                # Write order sheets first (these are needed for the summary tables)
                if mask_lil_o.any(): 
                    lil = self.df_ordini_all[mask_lil_o]
                    lil = lil[df_columns]
                    lil.to_excel(writer, sheet_name='Ordini LIL', index=False)
                
                if mask_agee_o.any():
                    agee = self.df_ordini_all[mask_agee_o]
                    agee = agee[df_columns]
                    agee.to_excel(writer, sheet_name='Ordini AGEE', index=False)

                # Write payment sheets
                mask_lil_p = self.pagamenti["Brand"] == "LIL Milan"
                mask_agee_p = self.pagamenti["Brand"] == "AGEE"

                if mask_lil_p.any():
                    for p in self.pagamenti["Metodo"].unique():
                        payment_name_lil = p.split()[0] + "_LIL"
                        filtered_df_lil = self.pagamenti[mask_lil_p & (self.pagamenti["Metodo"] == p)]

                        if not filtered_df_lil.empty:
                            matching_columns = next((cols for key, cols in pagamenti_columns.items() 
                                                     if key == p), None)
                
                            # Filter columns if match found
                            if len(matching_columns) > 0:
                                filtered_df_lil = filtered_df_lil[matching_columns]
                            filtered_df_lil.to_excel(writer, sheet_name=payment_name_lil, index=False)

                if mask_agee_p.any():
                    for p in self.pagamenti["Metodo"].unique():
                        payment_name_agee = p.split()[0] + "_AGEE"
                        filtered_df_agee = self.pagamenti[mask_agee_p & (self.pagamenti["Metodo"] == p)]
                        
                        if not filtered_df_agee.empty:
                            matching_columns = next((cols for key, cols in pagamenti_columns.items() 
                                                if key == p), None)

                            # Filter columns if match found
                            if len(matching_columns) > 0:
                                filtered_df_agee = filtered_df_agee[matching_columns]
                            filtered_df_agee.to_excel(writer, sheet_name=payment_name_agee, index=False)

            # Now create the summary tables after all required sheets exist
            self.create_summary_table()  # This creates 'Totale' sheet
            self.create_daily_summary_table()  # This creates 'Totale_daily' sheet

            # Finally reorder sheets
            workbook = load_workbook(self.filename)

            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

            self.reorder_sheets(workbook, sheet_order)
            workbook.save(self.filename)
                
            return self.filename

        except Exception as e:
            print(f"Error creating Excel file: {str(e)}")
            raise


    def reorder_sheets(self, workbook, sheet_order):
        """Helper function to reorder sheets in the workbook"""
        try:
            # Get the sheets that exist in the workbook
            existing_sheets = set(workbook.sheetnames)
            
            # Filter sheet_order to only include sheets that exist
            final_order = [sheet for sheet in sheet_order if sheet in existing_sheets]
            
            # Add any existing sheets that weren't in the order list at the end
            final_order.extend(sheet for sheet in existing_sheets if sheet not in final_order)
            
            # Reorder the sheets
            for i, sheet_name in enumerate(final_order):
                current_index = workbook.index(workbook[sheet_name])
                if current_index != i:  # Only move if needed
                    workbook.move_sheet(sheet_name, offset=i-current_index)
                    
        except Exception as e:
            print(f"Warning: Error while reordering sheets: {str(e)}")
            print("Continuing with original sheet order...")

    # Function to check and reformat
    def reformat_date(self, date_str):
        if pd.notna(date_str):
            date_str = date_str.strip().replace("/", "-")[:10]
        # Check if the string starts with "2024"
            if not date_str.startswith("2024"):
                # Split and rearrange as "YYYY-MM-DD"
                return "-".join(date_str.split("-")[::-1])
        return date_str  # Return as-is if already starts with "2024"

 
    # Method to create stats for each store location
    def create_location_stats(self, df, start_row, summary_sheet, store_name):
        location_stats = df.groupby('Location').agg({'Name': 'nunique',
                                                     'Lineitem quantity': 'sum'}).reset_index()
        stats_dict = location_stats.set_index('Location').to_dict()
        title_of_locations = df["Location"].unique()

        summary_sheet.merge_cells(f'H{start_row-1}:L{start_row-1}') 
        cell = summary_sheet[f'H{start_row-1}']
        cell.value = store_name
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        for idx, location_label in enumerate(title_of_locations, start=start_row):
            summary_sheet[f'H{idx}'] = location_label
            summary_sheet[f'I{idx}'] = (
                f'=SUMIFS(\'Ordini {store_name}\'!$L:$L, '
                f'\'Ordini {store_name}\'!$BB:$BB, "{location_label}")'
            )
            unique_orders = stats_dict['Name'].get(location_label, 0)
            items_quantity = stats_dict['Lineitem quantity'].get(location_label, 0)

            summary_sheet[f'J{idx}'] = unique_orders
            summary_sheet[f'K{idx}'] = items_quantity
            summary_sheet[f'L{idx}'] = f'=K{idx}/J{idx}'
            summary_sheet[f'L{idx}'].number_format = '0.00'

        summary_sheet[f'H{idx+1}'] = 'Totale'
        summary_sheet[f'H{idx+1}'].font = Font(bold=True)
        summary_sheet[f'I{idx+1}'] = f'=SUM(I{start_row}:I{idx})'
        summary_sheet[f'J{idx+1}'] = f'=SUM(J{start_row}:J{idx})'
        summary_sheet[f'K{idx+1}'] = f'=SUM(K{start_row}:K{idx})'
        summary_sheet[f'L{idx+1}'] = f'=K{idx+1}/J{idx+1}'
        summary_sheet[f'L{idx+1}'].number_format = '0.00'

        return start_row + len(title_of_locations) + 3

    # Method to create a summary table in Excel
    def create_summary_table(self):
        try:
            workbook = load_workbook(self.filename)
        except FileNotFoundError:
            workbook = Workbook()

        summary_sheet = workbook['Totale'] if 'Totale' in workbook.sheetnames else workbook.create_sheet('Totale')

        for row in summary_sheet.iter_rows(min_row=1, max_col=20, max_row=summary_sheet.max_row):
            for cell in row:
                cell.value = None

        bold_font = Font(bold=True)
        headers = {
            'A1': 'Payments', 'B1': 'LIL', 'C1': 'AGEE', 'D1': 'CHECK LIL', 'E1': 'CHECK AGEE'
        }
        for cell_position, value in headers.items():
            cell = summary_sheet[cell_position]
            cell.value = value
            cell.font = bold_font

        title_of_totals = {
            'Scalapay': 'J', 'Shopify': 'I', 'PayPal': 'H', 'Bonifico': 'H',
            'Qromo': 'F', 'Satispay': 'E', 'Cash': ''
        }

        row = 2
        for payment_label, payment_amount in title_of_totals.items():
            summary_sheet[f'A{row}'] = payment_label
            summary_sheet[f'B{row}'] = f'=SUMIFS(\'Ordini LIL\'!$L:$L, \'Ordini LIL\'!$AV:$AV, "*{payment_label}*")'
            summary_sheet[f'C{row}'] = f'=IFERROR(SUMIFS(\'Ordini AGEE\'!$L:$L, \'Ordini AGEE\'!$AV:$AV, "*{payment_label}*"), 0)'
            
            if payment_label == "Cash":
                summary_sheet[f'D{row}'] = '-'
                summary_sheet[f'E{row}'] = '-'
            
            elif payment_label == "Qromo":
                summary_sheet[f'D{row}'] = f'=IFERROR(SUM(\'{payment_label}_LIL\'!C:C) - SUM(\'{payment_label}_LIL\'!D:D), 0)'
                summary_sheet[f'E{row}'] = f'=IFERROR(SUM(\'{payment_label}_AGEE\'!C:C) - SUM(\'{payment_label}_AGEE\'!D:D), 0)'
            
            else:
                summary_sheet[f'D{row}'] = f'=IFERROR(SUM(\'{payment_label}_LIL\'!{payment_amount}:{payment_amount}), 0)'
                summary_sheet[f'E{row}'] = f'=IFERROR(SUM(\'{payment_label}_AGEE\'!{payment_amount}:{payment_amount}), 0)'

            row += 1

        summary_sheet[f'A{row}'] = 'Totale'
        summary_sheet[f'A{row}'].font = Font(bold=True)
        summary_sheet[f'B{row}'] = f'=SUM(B2:B{row-1})'
        summary_sheet[f'C{row}'] = f'=SUM(C2:C{row-1})'

        headers = {'H1': 'Locations', 'I1': 'Incasso', 'J1': 'Ordini', 'K1': 'Items', 'L1': 'Oggetti per ordine'}
        for cell_position, value in headers.items():
            cell = summary_sheet[cell_position]
            cell.value = value
            cell.font = bold_font

        exclude_strings = ["Luxury Pack", "Engraving", "E-gift", "Repair", "Whatever Tote", "Piercing Party", "LIL Bag"]
        df_ordini_gioielli = self.df_ordini_all[~self.df_ordini_all['Lineitem name'].str.contains('|'.join(exclude_strings), case=False, na=False)]
    
        df_lil = df_ordini_gioielli[df_ordini_gioielli['Brand'] == 'LIL Milan']
        df_agee = df_ordini_gioielli[df_ordini_gioielli['Brand'] == 'AGEE']

        start_row = 3
        start_row = self.create_location_stats(df_lil, start_row, summary_sheet, 'LIL')
        if len(df_agee) > 0:
            self.create_location_stats(df_agee, start_row, summary_sheet, 'AGEE')

        workbook.save(self.filename)

    # Method to create a daily summary table
    def create_daily_summary_table(self):
        try:
            workbook = load_workbook(self.filename)
        except FileNotFoundError:
            workbook = Workbook()

        daily_sheet = workbook['Totale_daily'] if 'Totale_daily' in workbook.sheetnames else workbook.create_sheet('Totale_daily')

        for row in daily_sheet.iter_rows(min_row=1, max_col=10, max_row=daily_sheet.max_row):
            for cell in row:
                cell.value = None

        bold_font = Font(bold=True)
        headers = {'A1': 'Giorno', 'B1': 'Paese', 'C1': 'Incasso LIL', 'D1': 'Incasso AGEE', 'E1': 'Incasso Totale'}
        for cell_position, value in headers.items():
            cell = daily_sheet[cell_position]
            cell.value = value
            cell.font = bold_font

        self.df_ordini_all['Giorno'] = self.df_ordini_all['Paid at'].apply(self.reformat_date)

        df_ordini_locations = self.df_ordini_all[self.df_ordini_all["Location"].isin(["Firgun House", "LIL House", "LIL House London"])]
        df_ordini_locations_lil = df_ordini_locations[df_ordini_locations["Brand"] == "LIL Milan"]
        df_ordini_locations_agee = df_ordini_locations[df_ordini_locations["Brand"] == "AGEE"]

        daily_country_totals_lil = df_ordini_locations_lil.groupby(['Giorno', 'Shipping Country'])['Total'].sum().reset_index().sort_values(['Giorno', 'Shipping Country']).rename(columns={'Shipping Country': 'Country', 'Total': 'Total_LIL'})
        daily_country_totals_agee = df_ordini_locations_agee.groupby(['Giorno', 'Shipping Country'])['Total'].sum().reset_index().sort_values(['Giorno', 'Shipping Country']).rename(columns={'Shipping Country': 'Country', 'Total': 'Total_AGEE'})

        daily_country_totals = pd.merge(daily_country_totals_lil, daily_country_totals_agee, on=["Giorno", "Country"], how="outer").fillna(0)

        for idx, row in enumerate(daily_country_totals.itertuples(), start=2):
            daily_sheet[f'A{idx}'] = row.Giorno
            daily_sheet[f'B{idx}'] = row.Country
            daily_sheet[f'C{idx}'] = row.Total_LIL
            daily_sheet[f'D{idx}'] = row.Total_AGEE
            daily_sheet[f'E{idx}'] = f'=C{idx}+D{idx}'

        #totals
        daily_sheet[f'B{idx+2}'] = 'Totale'
        daily_sheet[f'B{idx+2}'].font = Font(bold=True)
        daily_sheet[f'C{idx+2}'] =  f'=SUM(C2:C{idx})'
        daily_sheet[f'D{idx+2}'] = f'=SUM(D2:D{idx})'
        daily_sheet[f'E{idx+2}'] = f'=SUM(E2:E{idx})'

        workbook.save(self.filename)



